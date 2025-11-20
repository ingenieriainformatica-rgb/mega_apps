from odoo import api, fields, models, _
from odoo.exceptions import UserError

import base64
from io import BytesIO
from datetime import datetime, date, timedelta
import html as html_lib
import re


class UtilitiesExcelImportWizard(models.TransientModel):
    _name = 'utilities.excel.import.wizard'
    _description = 'Importar Servicios Publicos desde Excel'

    file = fields.Binary(string='Archivo Excel', required=True)
    filename = fields.Char()
    sheet_name = fields.Char(string='Hoja (opcional)')
    simulate = fields.Boolean(string='Simular (no crear)', default=False)
    summary = fields.Text(readonly=True)
    summary_html = fields.Html(string='Resumen', readonly=True, sanitize=False)

    # -------------------------
    # Helpers
    # -------------------------
    def _norm(self, s):
        s = (s or '').strip().lower()
        return ' '.join(s.split())

    def _to_date(self, v):
        if not v:
            return None
        if isinstance(v, (date, datetime)):
            return v.date() if isinstance(v, datetime) else v
        vs = str(v).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                return datetime.strptime(vs, fmt).date()
            except Exception:
                continue
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(v))).date()
        except Exception:
            return None

    def _read_xlsx(self, stream, sheet_name=None):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise UserError(_("Falta el paquete Python 'openpyxl' para leer archivos .xlsx. Por favor instálarlo."))

        wb = load_workbook(stream, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [self._norm(h or '') for h in rows[0]]
        res = []
        for r in rows[1:]:
            if not r:
                continue
            data = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = r[i] if i < len(r) else None
                data[h] = val
            if all(v in (None, '', 0) for v in data.values()):
                continue
            res.append(data)
        return headers, res

    def _read_file_rows(self):
        if not self.file:
            raise UserError(_("Debe adjuntar un archivo."))
        data = base64.b64decode(self.file)
        fname = (self.filename or '').lower()

        if fname.endswith('.xlsx') or not fname:
            return self._read_xlsx(BytesIO(data), sheet_name=self.sheet_name)

        if fname.endswith('.csv'):
            import csv
            try:
                text = data.decode('utf-8')
            except Exception:
                text = data.decode('latin-1')
            reader = csv.DictReader(text.splitlines())
            headers = [self._norm(h or '') for h in (reader.fieldnames or [])]
            rows = [{self._norm(k): v for k, v in (row or {}).items()} for row in reader]
            return headers, rows

        raise UserError(_("Formato no soportado: %s (use .xlsx o .csv)" % (self.filename or 'sin nombre')))

    def _get_meter(self, kind):
        Meter = self.env['x.meter'].sudo()
        k = self._norm(kind)
        if k.startswith('agua') or k.startswith('water'):
            m = Meter.search([('name', 'ilike', 'agua')], limit=1)
            if not m:
                m = Meter.search([('name', 'ilike', 'water')], limit=1)
            if not m:
                m = Meter.create({'name': 'Agua', 'x_price': 0.0})
            return m
        if k.startswith('ener') or k.startswith('elec'):
            m = Meter.search([('name', 'ilike', 'ener')], limit=1)
            if not m:
                m = Meter.search([('name', 'ilike', 'elec')], limit=1)
            if not m:
                m = Meter.create({'name': 'Energía', 'x_price': 0.0})
            return m
        m = Meter.search([('name', 'ilike', kind)], limit=1)
        return m or Meter.create({'name': kind or 'Medidor', 'x_price': 0.0})

    def _find_property(self, id_propiedad=None, dir_text=None):
        Acc = self.env['account.analytic.account'].sudo()
        domain = [('x_is_property', '=', True)]
        if id_propiedad:
            s = str(id_propiedad).strip()
            digits = re.sub(r'\D', '', s)
            if digits:
                candidates = Acc.search(domain + [('x_cons_inm', 'ilike', digits)], limit=100)
                for c in candidates:
                    cv = re.sub(r'\D', '', (c.x_cons_inm or ''))
                    if cv == digits:
                        return c
            prop = Acc.search(domain + [('x_cons_inm', '=', s)], limit=1)
            if prop:
                return prop
            return Acc.browse()
        if dir_text:
            dt = (dir_text or '').strip()
            prop = Acc.search(domain + ['|', '|',
                                   ('name', 'ilike', dt),
                                   ('x_property_geolocation', 'ilike', dt),
                                   ('x_property_address', 'ilike', dt)
                               ], limit=1)
            if prop:
                return prop
        return Acc.browse()

    def _row_to_values(self, row):
        r = {self._norm(k): v for k, v in (row or {}).items()}
        id_prop = r.get('id_propiedad') or r.get('id')
        dir_text = r.get('dir') or r.get('direccion')
        apto = r.get('apto') or r.get('apartment') or ''
        fecha = self._to_date(r.get('fecha_lectura') or r.get('fecha'))
        agua = r.get('agua')
        energia = r.get('energia') or r.get('enerigia') or r.get('electricidad') or r.get('electricity')
        return id_prop, dir_text, apto, fecha, agua, energia

    def action_import(self):
        self.ensure_one()
        headers, rows = self._read_file_rows()
        hdr = set(headers)
        required_alt = [
            ('ID', {'id', 'id_propiedad'}),
            ('dir', {'dir', 'direccion'}),
            ('apto', {'apto', 'apt', 'apartment'}),
            ('fecha_lectura', {'fecha_lectura', 'fecha', 'fecha lectura', 'date'}),
            ('agua', {'agua', 'water'}),
            ('energia', {'energia', 'enerigia', 'energía', 'electricidad', 'electricity'}),
        ]
        missing = []
        for canon, alts in required_alt:
            if not any(a in hdr for a in alts):
                missing.append(canon)
        if missing:
            raise UserError(_("El archivo no contiene las columnas requeridas: %s") % ", ".join(missing))
        if not rows:
            self.summary = _("No se encontraron filas para importar.")
            return {'type': 'ir.actions.act_window_close'}

        meter_agua = self._get_meter('agua')
        meter_ener = self._get_meter('energia')

        created = 0
        updated = 0
        errors = []

        MR = self.env['x.meter.reading'].sudo()

        for idx, row in enumerate(rows, start=2):
            id_prop, dir_text, apto, fecha, agua, energia = self._row_to_values(row)

            if not (id_prop or dir_text):
                errors.append(_(f"Fila {idx}: Falta 'id_propiedad' o 'dir'."))
                continue
            if not fecha:
                errors.append(_(f"Fila {idx}: 'fecha_lectura' invalida o vacia."))
                continue

            prop = None
            try:
                id_digits = re.sub(r'\D', '', str(id_prop or '').strip())
                rec_id = int(id_digits) if id_digits else None
            except Exception:
                rec_id = None
            if rec_id:
                acc = self.env['account.analytic.account'].sudo().browse(rec_id)
                if acc and acc.exists() and getattr(acc, 'x_is_property', False):
                    prop = acc
            if not prop:
                prop = self._find_property(id_propiedad=str(id_prop or '').strip(), dir_text=dir_text)
            if not prop:
                errors.append(_(f"Fila {idx}: Propiedad no encontrada (id={id_prop} dir='{dir_text}')."))
                continue

            desc = None
            if dir_text or apto:
                desc = f"{dir_text or ''} - Apto {apto}".strip(' -')

            def _upsert(meter, qty, unit_cost):
                nonlocal created, updated
                if qty in (None, ''):
                    return
                try:
                    q = float(qty)
                except Exception:
                    errors.append(_(f"Fila {idx}: Lectura no numerica para medidor '{meter.name}'."))
                    return
                # 1) Solo una importacion por mes por propiedad+medidor
                month_begin = date(fecha.year, fecha.month, 1)
                if fecha.month == 12:
                    month_end = date(fecha.year, 12, 31)
                else:
                    month_end = date(fecha.year, fecha.month + 1, 1) - timedelta(days=1)
                existing_month = MR.search([
                    ('x_account_analytic_account_id', '=', prop.id),
                    ('x_meter_id', '=', meter.id),
                    ('x_date', '>=', month_begin),
                    ('x_date', '<=', month_end),
                ], limit=1)
                if existing_month:
                    errors.append(_(f"Fila {idx}: Ya existe una lectura para '{meter.name}' en {fecha.strftime('%Y-%m')} en esta propiedad."))
                    return
                # 2) La lectura debe ser >= al valor del mes anterior
                last_prev_month = MR.search([
                    ('x_account_analytic_account_id', '=', prop.id),
                    ('x_meter_id', '=', meter.id),
                    ('x_date', '<', month_begin),
                ], order='x_date desc, id desc', limit=1)
                if last_prev_month:
                    try:
                        prev_q = float(last_prev_month.x_quantity or 0.0)
                    except Exception:
                        prev_q = 0.0
                    if q < prev_q:
                        errors.append(_(f"Fila {idx}: La lectura {q} es menor que la lectura del mes anterior {prev_q} para el medidor '{meter.name}'."))
                        return
                # 3) Mantener verificacion general de orden temporal (no retroceder)
                last_any = MR.search([
                    ('x_account_analytic_account_id', '=', prop.id),
                    ('x_meter_id', '=', meter.id),
                ], order='x_date desc, id desc', limit=1)
                if last_any and fecha <= last_any.x_date:
                    errors.append(_(f"Fila {idx}: La fecha de lectura ({fecha}) debe ser posterior a la ultima guardada ({last_any.x_date}) para el medidor '{meter.name}'."))
                    return
                if self.simulate:
                    created += 1
                    return
                MR.create({
                    'x_account_analytic_account_id': prop.id,
                    'x_meter_id': meter.id,
                    'x_date': fecha,
                    'x_quantity': q,
                    'x_description': desc,
                    'x_unit_cost': unit_cost,
                })
                created += 1

            _upsert(meter_agua, agua, getattr(prop.company_id, 'utility_price_water', 0.0) or 0.0)
            _upsert(meter_ener, energia, getattr(prop.company_id, 'utility_price_energy', 0.0) or 0.0)

        summary_lines = [
            _("Importacion completada."),
            _("Creados: %s") % created,
            _("Actualizados: %s") % updated,
            _("Errores: %s") % len(errors),
        ]
        if errors:
            summary_lines.append("\n".join(errors[:50]))
        self.summary = "\n".join(summary_lines)

        ok = len(errors) == 0
        badges = (
            f"<span class='badge bg-success me-2'>Creados: {created}</span>"
            f"<span class='badge bg-primary me-2'>Actualizados: {updated}</span>"
            f"<span class='badge {'bg-danger' if errors else 'bg-success'}'>Errores: {len(errors)}</span>"
        )
        header = (
            f"<div class='alert {'alert-success' if ok else 'alert-warning'}' role='alert'>"
            f"<strong>Importacion completada.</strong> {badges}"
            f"</div>"
        )
        errors_html = ""
        if errors:
            rows = "".join(
                f"<tr><td style='width:40px'>{i}</td><td>{html_lib.escape(e)}</td></tr>"
                for i, e in enumerate(errors[:50], start=1)
            )
            errors_html = (
                "<div class='mt-2'><strong>Detalles de errores (max 50):</strong></div>"
                f"<table class='table table-sm table-striped'><thead><tr><th>#</th><th>Detalle</th></tr></thead><tbody>{rows}</tbody></table>"
            )
        self.summary_html = header + errors_html
        try:
            self.env.user.notify_success(message=_('Archivo importado correctamente.'))
        except Exception:
            pass

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

