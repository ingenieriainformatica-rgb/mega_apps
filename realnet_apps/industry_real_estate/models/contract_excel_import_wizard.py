# -*- coding: utf-8 -*-
from odoo import api, fields, models, _, tools
from odoo.exceptions import UserError
from odoo.tools.float_utils import float_is_zero
from datetime import datetime
import base64
from io import BytesIO
import logging
import re, unicodedata
import unicodedata  # asegúrate de tenerlo
from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)
IVA_RATE = 0.19


class ContractExcelImportWizard(models.TransientModel):
    _name = 'contract.excel.import.wizard'
    _description = 'Importar contratos ECOERP desde Excel'

    file = fields.Binary(string='Archivo Excel', required=True)
    filename = fields.Char()
    sheet_name = fields.Char(string='Hoja (opcional)')
    simulate = fields.Boolean(string='Simular (no crear)', default=False)
    summary = fields.Text(readonly=True)
    
    def _sanitize_vat(self, v):
        return re.sub(r'\D', '', self._s(v))

    def _search_partner_by_vat_digits(self, raw_vat):
        """
        Busca un partner por coincidencia EXACTA de dígitos de VAT,
        considerando archivados (active_test=False).
        """
        Partner = self.env['res.partner'].with_context(active_test=False).sudo()
        digits = self._sanitize_vat(raw_vat)
        if not digits:
            return Partner.browse()
        # Trae candidatos por ilike y filtra por igualdad de dígitos
        candidates = Partner.search([('vat', 'ilike', digits)], limit=120)
        for p in candidates:
            if self._sanitize_vat(p.vat) == digits:
                return p
        return Partner.browse()

    
    def _is_owners_only_row(self, row, headers_rev, C):
        """Segunda plantilla: sin canon y sin datos de inquilino (con headers mapeados)."""
        canon_raw = self._get(row, headers_rev, C['canon'])
        canon_num = self._to_float(canon_raw)
        arr_nom   = self._s(self._get(row, headers_rev, C['arr_nom']))
        arr_vat   = self._s(self._get(row, headers_rev, C['arr_vat']))
        return not canon_num and not arr_nom and not arr_vat
    
    def _iter_owner_blocks(self, row, headers_rev):
        """
        Devuelve lista de dicts:
        { owner_doc_raw, owner_vat, owner_name, owner_pct, benef_doc_raw, benef_vat, benef_name }

        PRIORIDAD DE %:
        - Propietario 1  -> 'Participacion_Prop'    (si hay valor)
        - Propietario 2  -> 'Participacion_Prop2'   (si hay valor)
        - Si no hay en esos campos, cae a:
        'porcentaje participacion PROP X' del bloque correspondiente.
        """
        blocks = []
        H = lambda name: self._get(row, headers_rev, name)

        # Leer los dos campos “globales” del Excel (pueden venir vacíos)
        p1_global = self._to_float(H('Participacion_Prop'))      # % para OWNER 1
        p2_global = self._to_float(H('Participacion_Prop2'))     # % para OWNER 2

        for idx in range(1, 6):
            if idx == 1:
                odoc = self._s(H('Tipo doc prop 1'))
                ovat = self._s(H('Cédula/Nit Propietario 1'))
                onam = self._s(H('Nombre Propietario 1'))
                # % del bloque (fallback)
                opct_block = self._to_float(H('porcentaje participacion PROP 1'))
                pcte = self._to_float(H('Porcentaje_comision prop 1'))
                # prioridad: Participacion_Prop si existe
                opct = p1_global if p1_global else opct_block

                bdoc = self._s(H('Tipo doc BENEF PROP 1'))
                bvat = self._s(H('Cédula/Nit Beneficiario BENEF PROP 1'))
                bnam = self._s(H('Nombre Beneficiario BENEF PROP 1')) 

            elif idx == 2:
                odoc = self._s(H('Tipo doc  prop 2'))
                ovat = self._s(H('Cédula/Nit Propietario 2'))
                onam = self._s(H('Nombre Propietario 2'))
                opct_block = self._to_float(H('porcentaje participacion PROP 2'))
                pcte = self._to_float(H('Porcentaje_comision prop 2'))
                # prioridad: Participacion_Prop2 si existe
                opct = p2_global if p2_global else opct_block

                bdoc = self._s(H('Tipo doc benef prop 2'))
                bvat = self._s(H('Cédula/Nit Beneficiario BENEF PROP 2'))
                bnam = self._s(H('Nombre Beneficiario BENEF PROP 2'))    

            elif idx == 3:
                odoc = self._s(H('Tipo doc prop 3'))
                ovat = self._s(H('Cédula/Nit Propietario 3'))
                onam = self._s(H('Nombre Propietario 3'))
                opct = self._to_float(H('porcentaje participacion PROP 3'))
                pcte = self._to_float(H('Porcentaje_comision prop 3'))

                bdoc = self._s(H('Tipo doc bener prop 3'))
                bvat = self._s(H('Cédula/Nit Beneficiario BENEF PROP 3'))
                bnam = self._s(H('Nombre Beneficiario BENEF PROP 3'))

            elif idx == 4:
                odoc = self._s(H('Tipo doc prop 4'))
                ovat = self._s(H('Cédula/Nit Propietario 4'))
                onam = self._s(H('Nombre Propietario 4'))
                opct = self._to_float(H('porcentaje participacion PROP 4'))
                pcte = self._to_float(H('Porcentaje_comision prop 4'))

                bdoc = self._s(H('Tipo doc benef prop 4'))
                bvat = self._s(H('Cédula/Nit Beneficiario BENEF PROP 4'))
                bnam = self._s(H('Nombre Beneficiario BENEF PROP 4'))

            else:  # idx == 5
                odoc = self._s(H('Tipo doc prop 5'))
                ovat = self._s(H('Cédula/Nit Propietario 5'))
                onam = self._s(H('Nombre Propietario 5'))
                opct = self._to_float(H('porcentaje participacion PROP 5'))
                pcte = self._to_float(H('Porcentaje_comision prop 5'))

                bdoc = self._s(H('Tipo doc benef prop 5'))
                bvat = self._s(H('Cédula/Nit Beneficiario BENEF PROP 5'))
                bnam = self._s(H('Nombre Beneficiario BENEF PROP 5'))

            # bloque válido SOLO si hay documento del propietario
            if not ovat:
                continue

            blocks.append({
                'owner_doc_raw': odoc,
                'owner_vat': ovat,
                'owner_name': onam,
                'owner_pct': opct or 0.0,
                'benef_doc_raw': bdoc,
                'benef_vat': bvat,
                'benef_name': bnam,
                'comis_espe':pcte,
            })

        return blocks

    
    def _existing_owner_ids(self, prop):
        model_name, prop_field, _pf = self._resolve_owner_line_model()
        if not (model_name and prop_field):
            return set()
        M = self.env[model_name].sudo()
        owners = M.search([(prop_field, '=', prop.id)])
        return set(owners.mapped('owner_id').ids)


    def _sum_owner_percent(self, prop):
        model_name, prop_field, percent_field = self._resolve_owner_line_model()
        if not (model_name and prop_field and percent_field):
            return 0.0
        M = self.env[model_name].sudo()
        total = 0.0
        for l in M.search([(prop_field, '=', prop.id)]):
            try:
                total += float(getattr(l, percent_field) or 0.0)
            except Exception:
                pass
        return total


    def _ensure_owner_only_create(self, prop, owner, percent=0.0,comis_espe=0.0):
        """Crea la línea de owner SOLO si no existe; nunca actualiza."""
        model_name, prop_field, percent_field = self._resolve_owner_line_model()
        if not model_name:
            return False
        M = self.env[model_name].sudo()
        exists = M.search([(prop_field, '=', prop.id), ('owner_id', '=', owner.id)], limit=1)
        if exists:
            return exists  # no tocar

        vals = {prop_field: prop.id, 'owner_id': owner.id,'comision_personalizada':comis_espe}
        if percent_field:
            _logger.info(" _!!!! %s",percent)            
            vals[percent_field] = percent or 0.0
        if 'company_id' in M._fields and prop.company_id:
            vals['company_id'] = prop.company_id.id
        line = M.create(vals)
        # seguridad: sin representante
        if 'representative_id' in line._fields and line.representative_id:
            line.sudo().write({'representative_id': False})
        # 🔹 Propietario NO es beneficiario por defecto
        if 'is_main_payee' in line._fields and line.is_main_payee:
            line.sudo().write({'is_main_payee': False})
        return line


    def _ensure_benef_only_create(self, prop, owner_line, beneficiary, percent=0.0):
        """Crea beneficiario SOLO si no existe (por propiedad+beneficiario),
        y bloquea cuando el beneficiario coincide con un propietario.
        """
        model_name, parent_field, prop_field = self._resolve_benef_line_model()
        if not (model_name and parent_field and prop_field):
            return False

        M = self.env[model_name].sudo()

        # --- Resolver property id de forma robusta ---
        prop_id = False
        if owner_line and (prop_field in owner_line._fields) and getattr(owner_line, prop_field):
            prop_id = getattr(owner_line, prop_field).id
        elif prop:
            prop_id = prop.id
        if not prop_id:
            return False

        # ===================== GUARDAS ANTI-DUPLICIDAD PROP/BNF =====================

        # a) Si el beneficiario es el MISMO partner del owner_line, bloquear
        if owner_line and getattr(owner_line, 'owner_id', False) and owner_line.owner_id.id == beneficiary.id:
            propiedad = prop.display_name or prop.name
            raise ValidationError(_(
                "El propietario y el beneficiario no pueden ser la misma persona:"
                "\nPropiedad %(propiedad)s tiene a %(persona)s como propietario y beneficiario."
            ) % {
                'propiedad': propiedad,
                'persona': owner_line.owner_id.display_name or owner_line.owner_id.name,
            })

        # ===================== FIN GUARDAS =====================

        # Evitar duplicado de línea de beneficiario (misma propiedad + mismo partner)
        exists = M.search([
            (prop_field, '=', prop_id),
            ('owner_id', '=', beneficiary.id),
            ('parent_owner_line_id', '=', owner_line.id),
        ], limit=1)
        # exists = M.search([(prop_field, '=', prop_id), ('owner_id', '=', beneficiary.id)], limit=1)
        if exists:
            updates = {}
            if 'beneficial_porcentage' in M._fields and float(exists.beneficial_porcentage or 0.0) != float(percent or 0.0):
                _logger.info(" !!!! %s",percent)
                updates['beneficial_porcentage'] = percent or 0.0
            # Beneficiario SIEMPRE marcado
            if 'is_main_payee' in M._fields and not exists.is_main_payee:
                updates['is_main_payee'] = True
            if updates:
                exists.sudo().write(updates)
            return exists

        # Crear nueva línea de beneficiario
        vals = {
            prop_field: prop_id,
            parent_field: owner_line.id if owner_line else False,
            'owner_id': beneficiary.id,
            'is_main_payee': True,
            'beneficial_porcentage': percent or 0.0,
        }
        if 'company_id' in M._fields and prop and prop.company_id:
            vals['company_id'] = prop.company_id.id

        return M.create(vals)


    def _owners_only_branch(self, *, i, prop, blocks, lines_log, so_name=None, **_ignored):
        """
        Crea SOLO los owners/beneficiarios que no existan.
        No actualiza nada de los existentes.
        Valida que la suma no supere 100%.
        """
        if not prop:
            lines_log.append(f"Fila {i}: sin propiedad (Cons_Inm) → omitido.")
            return

        # 1) Resolver partners y armar “nuevos” candidatos
        new_blocks, sum_new = [], 0.0
        for b in blocks:
            try:
                owner_ident_id = self._get_latam_ident_type_id(b['owner_doc_raw'])
                owner_do_nit = bool(owner_ident_id and self._is_nit_type(owner_ident_id))
                owner = self._upsert_partner_by_vat(
                    b['owner_name'], b['owner_vat'],
                    ident_type_id=owner_ident_id, do_nit=owner_do_nit, overwrite=False
                )
                if not owner:
                    lines_log.append(f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: propietario inválido ({b['owner_name']} / {b['owner_vat']}) → omitido bloque.")
                    continue

                beneficiary = None
                if b.get('benef_vat') and b.get('benef_name'):
                    benef_ident_id = self._get_latam_ident_type_id(b.get('benef_doc_raw'))
                    beneficiary = self._get_or_create_partner_by_vat(b['benef_vat'], b['benef_name'], ident_type_id=benef_ident_id, do_nit=False)

                new_blocks.append({'owner': owner, 'beneficiary': beneficiary, 'pct': float(b.get('owner_pct') or 0.0), 'comis_espe': float(b.get('comis_espe') or 0.0)})
            except Exception as e:
                lines_log.append(f"Fila {i}: error resolviendo partners del bloque → {e}")

        if not new_blocks:
            lines_log.append(f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: sin bloques válidos → omitido.")
            return

        # 2) Tope 100% SOLO con los realmente nuevos
        existing_owner_ids = self._existing_owner_ids(prop)
        current_sum = self._sum_owner_percent(prop)
        sum_new = sum(blk['pct'] for blk in new_blocks if blk['owner'].id not in existing_owner_ids)
        if current_sum + sum_new > 100.0 + 1e-6:
            lines_log.append(
                f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: "
                f"total {current_sum:.2f}% + nuevos {sum_new:.2f}% supera 100% → propiedad omitida completa."
            )
            return

        # 3) Crear solo los que no existan; existentes → omitir (sin actualizar)
        for blk in new_blocks:
            owner = blk['owner']
            pct   = blk['pct'] or 0.0
            benef = blk['beneficiary']
            comis_espe = blk['comis_espe']

            if owner.id in existing_owner_ids:
                lines_log.append(
                    f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: "
                    f"propietario {owner.display_name} ya existía → omitido."
                )
                # Beneficiario: solo crear si no existe
                model_name, prop_field, _pf = self._resolve_owner_line_model()
                owner_M = self.env[model_name].sudo() if model_name else False
                ol = owner_M.search([(prop_field, '=', prop.id), ('owner_id', '=', owner.id)], limit=1) if owner_M else False
                if benef and ol:
                    bl = self._ensure_benef_only_create(prop, ol, benef, percent=pct)
                    if bl:
                        lines_log.append(
                            f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: "
                            f"beneficiario {benef.display_name} creado ({pct:.2f}%)."
                        )
                continue

            # Crear owner
            owner_line = self._ensure_owner_only_create(prop, owner, percent=pct,comis_espe=comis_espe)
            if owner_line:
                lines_log.append(
                    f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: "
                    f"propietario {owner.display_name} creado ({pct:.2f}%)."
                )
                if benef:
                    bl = self._ensure_benef_only_create(prop, owner_line, benef, percent=pct)
                    if bl:
                        lines_log.append(
                            f"Fila {i} (Contrato {so_name or '-'}) Prop {prop.display_name}: "
                            f"beneficiario {benef.display_name} creado ({pct:.2f}%)."
                        )



    # -------------------------------
    # Helpers de resolución de modelo
    # -------------------------------
    def _resolve_owner_line_model(self):
        candidates = [
            'x_property_owner_line',
            'x_owner_property_line',
            'x_property_owners_line',
            'account.analytic.account.owner.line',
        ]
        for name in candidates:
            if name not in self.env:
                continue
            M = self.env[name]
            prop_field = 'analytic_account_id' if 'analytic_account_id' in M._fields else (
                'x_property_id' if 'x_property_id' in M._fields else False
            )
            if not prop_field or 'owner_id' not in M._fields:
                continue
            percent_field = next((pf for pf in ('participation_percent', 'x_percentage', 'percentage')
                                if pf in M._fields), None)
            # Si no hay campo de % en este modelo, mejor devuelve False para no prometer algo que no podrás escribir
            return name, prop_field, percent_field or False
        return False, False, False


    def _resolve_benef_line_model(self):
        for name in ('x_property_beneficiary_line', 'x_property_beneficiaries_line', 'account.analytic.account.owner.line'):
            if name not in self.env:
                continue
            M = self.env[name]
            parent_field = 'parent_owner_line_id' if 'parent_owner_line_id' in M._fields else False
            prop_field = 'analytic_account_id' if 'analytic_account_id' in M._fields else (
                'x_property_id' if 'x_property_id' in M._fields else False
            )
            if parent_field and prop_field and 'owner_id' in M._fields:
                # _logger.info("\n\n RESULTS::: (%s,%s,%s)",name, parent_field, prop_field)
                return name, parent_field, prop_field
        return False, False, False
    
        
    # --- Utils ---
    @api.model
    def _align_company(self, rec, comp_id):
        """Fuerza company_id si el modelo lo tiene y está distinto."""
        if rec and 'company_id' in rec._fields:
            if not rec.company_id or rec.company_id.id != comp_id:
                rec.sudo().write({'company_id': comp_id})

    @api.model
    def _clear_representative(self, rec):
        """Limpia representative_id si existe para evitar constraint."""
        if rec and 'representative_id' in rec._fields and rec.representative_id:
            rec.sudo().write({'representative_id': False})

    # -----------------------------------------
    # Crear/obtener línea de PROPIETARIO (owner)
    # -----------------------------------------
    @api.model
    def _link_owner_line(self, prop, owner, comp_id, percent_pro=100.0, main_payee=False, notes=None, percent_ben=0.0):
        if not (prop and owner):
            return False

        model_name, prop_field, percent_field = self._resolve_owner_line_model()
        if not model_name:
            return False

        M = self.env[model_name]
        dom = [(prop_field, '=', prop.id), ('owner_id', '=', owner.id)]
        existing = M.with_context(active_test=False).search(dom, limit=1)

        def _ensure_unique_main_payee(line):
            if 'is_main_payee' not in M._fields or not line.is_main_payee:
                return
            # desmarca otros principales de la misma propiedad
            others = M.search([(prop_field, '=', getattr(line, prop_field).id),
                            ('id', '!=', line.id),
                            ('is_main_payee', '=', True)])
            if others:
                others.sudo().write({'is_main_payee': False})

        if existing:
            vals_upd = {}
            if 'active' in existing._fields and existing.active is False:
                vals_upd['active'] = True
            if 'participation_percent' and (existing['participation_percent'] or 0.0) != (percent_pro or 0.0):
                _logger.info(" _!_!!! %s %s",percent_pro, owner.name)
                vals_upd['participation_percent'] = percent_pro or 0.0
            if 'beneficial_porcentage' and (existing['beneficial_porcentage'] or 0.0) != (percent_ben or 0.0):
                _logger.info(" _!_!!! %s",percent_ben)
                vals_upd['beneficial_porcentage'] = percent_ben or 0.0
            if 'is_main_payee' in M._fields and bool(existing.is_main_payee) != bool(main_payee):
                vals_upd['is_main_payee'] = bool(main_payee)
            if 'notes' in M._fields and (existing.notes or '') != (notes or ''):
                vals_upd['notes'] = notes or ''
            if vals_upd:
                existing.with_company(comp_id).sudo().write(vals_upd)
                _ensure_unique_main_payee(existing)
            return existing

        vals = {prop_field: prop.id, 'owner_id': owner.id}
        if percent_field:
            _logger.info(" _!!!E! %s",percent)
            vals[percent_field] = percent or 0.0
        if 'is_main_payee' in M._fields:
            vals['is_main_payee'] = bool(main_payee)
        if 'notes' in M._fields:
            vals['notes'] = notes or ''
        if 'company_id' in M._fields and comp_id:
            vals['company_id'] = comp_id

        line = M.with_company(comp_id).with_context(
            skip_lessor_contact_company_check=True, default_representative_id=False, force_company=comp_id
        ).sudo().create(vals)

        if 'representative_id' in line._fields and line.representative_id:
            line.sudo().write({'representative_id': False})

        _ensure_unique_main_payee(line)
        return line



    # ------------------------------------------------
    # Crear/obtener línea de BENEFICIARIO (beneficiary)
    # ------------------------------------------------

    @api.model
    def _link_benef_line(self, prop, owner_line, beneficiary, comp_id, percent=100.0, main_payee=False, notes=None):
        if not (owner_line and beneficiary):
            return False

        # 🔹 Caso especial: beneficiario == propietario → usar la misma línea y ENCENDER check
        if getattr(owner_line, 'owner_id', False) and owner_line.owner_id.id == beneficiary.id:
            vals_upd = {}
            if 'beneficial_porcentage' in owner_line._fields and (owner_line.beneficial_porcentage or 0.0) != (percent or 0.0):
                _logger.info(" _!!..!! %s",percent)
                vals_upd['beneficial_porcentage'] = percent or 0.0
            if 'notes' in owner_line._fields and (owner_line.notes or '') != (notes or ''):
                vals_upd['notes'] = notes or ''
            # 🔹 Beneficiario SIEMPRE marcado
            if 'is_main_payee' in owner_line._fields and not owner_line.is_main_payee:
                vals_upd['is_main_payee'] = True
            if vals_upd:
                owner_line.with_company(comp_id).sudo().write(vals_upd)
                if owner_line.is_main_payee and 'analytic_account_id' in owner_line._fields:
                    M = owner_line.__class__
                    prop_field = 'analytic_account_id' if 'analytic_account_id' in M._fields else 'x_property_id'
                    others = M.search([(prop_field, '=', getattr(owner_line, prop_field).id),
                                    ('id', '!=', owner_line.id),
                                    ('is_main_payee', '=', True)])
                    if others:
                        others.sudo().write({'is_main_payee': False})
            return owner_line

        model_name, parent_field, prop_field = self._resolve_benef_line_model()
        if not (model_name and parent_field and prop_field):
            return False
        M = self.env[model_name]

        # resolver propiedad
        prop_id = getattr(owner_line, prop_field).id if (prop_field in owner_line._fields and getattr(owner_line, prop_field)) else (prop.id if prop else False)
        prop_id = False
        if prop_field in owner_line._fields and getattr(owner_line, prop_field):
            prop_id = getattr(owner_line, prop_field).id
        elif prop and hasattr(prop, 'id'):
            prop_id = prop.id
        if not prop_id:
            return False

        existing = M.with_context(active_test=False).search([
            (prop_field, '=', prop_id), ('owner_id', '=', beneficiary.id)
        ], limit=1)
        
        def _ensure_unique_main_payee(line):
            if 'is_main_payee' not in M._fields or not line.is_main_payee:
                return
            others = M.search([(prop_field, '=', prop_id),
                            ('id', '!=', line.id),
                            ('is_main_payee', '=', True)])
            if others:
                others.sudo().write({'is_main_payee': False})

        if existing:
            vals_upd = {}
            if 'active' in existing._fields and existing.active is False:
                vals_upd['active'] = True
            if parent_field in existing._fields and existing[parent_field].id != owner_line.id:
                vals_upd[parent_field] = owner_line.id
            if 'beneficial_porcentage' in M._fields and (existing.beneficial_porcentage or 0.0) != (percent or 0.0):
                vals_upd['beneficial_porcentage'] = percent or 0.0
            if 'notes' in M._fields and (existing.notes or '') != (notes or ''):
                vals_upd['notes'] = notes or ''
            # 🔹 Beneficiario SIEMPRE marcado
            if 'is_main_payee' in M._fields and not existing.is_main_payee:
                vals_upd['is_main_payee'] = True
            if vals_upd:
                existing.with_company(comp_id).sudo().write(vals_upd)
            return existing

        vals = {
            prop_field: prop_id,
            parent_field: owner_line.id,
            'owner_id': beneficiary.id,
        }
        if 'beneficial_porcentage' in M._fields:
            vals['beneficial_porcentage'] = percent or 0.0
        # 🔹 Beneficiario SIEMPRE marcado
        if 'is_main_payee' in M._fields:
            vals['is_main_payee'] = True
        if 'notes' in M._fields:
            vals['notes'] = notes or ''
        if 'company_id' in M._fields and comp_id:
            vals['company_id'] = comp_id

        line = M.with_company(comp_id).with_context(
            skip_lessor_contact_company_check=True, default_representative_id=False, force_company=comp_id
        ).sudo().create(vals)

        # Asegura único principal
        _ensure_unique_main_payee(line)
        return line



    # -------------------------
    # Helpers mínimos (solo los usados)
    # -------------------------
    def _ensure_building_by_name_address(self, urb_name, addr):
        """
        Upsert de edificio por (x_name, x_address).
        Devuelve un recordset x_buildings (0..1).
        """
        name = (urb_name or '').strip()
        address = (addr or '').strip()
        if not name or not address:
            return self.env['x_buildings']  # vacío

        # B = self.env['x_buildings'].sudo().with_context(active_test=False)
        B = self.env['x_buildings'].sudo()

        # 1) Buscar exacto por la dupla (lo más rápido/fiable con el constraint)
        b = B.search([('x_name', '=', name), ('x_address', '=', address)], limit=1)
        if b:
            # si estaba archivado, lo reactivamos
            if hasattr(b, 'active') and not b.active:
                b.write({'active': True})
            return b

        # 2) Crear (respetará el constraint unique)
        vals = {
            'x_name': name,
            'x_address': address,
        }
        return B.create(vals)

    def _upsert_partner_by_vat(self, name, vat, ident_type_id=None, *, do_nit=False, overwrite=False):
        """
        Usa tu _find_existing_partner_by_vat para deduplicar por documento (tolerante a formato).
        - Si existe: completa solo campos vacíos (o pisa si overwrite=True) y normaliza NIT.
        - Si no existe: crea con NIT normalizado (si aplica).
        """
        Partner = self.env['res.partner'].sudo()
        name_clean = ' '.join(self._s(name).split())
        vat_raw = self._s(vat)
        
        # 🔒 Guardia: si ya existe por dígitos, reusar
        existing = self._search_partner_by_vat_digits(vat_raw)
        if existing:
            to_write = {}
            if ident_type_id and not existing.l10n_latam_identification_type_id:
                to_write['l10n_latam_identification_type_id'] = ident_type_id
            # Solo completa nombre si está vacío o si overwrite=True
            if name_clean and (overwrite or not existing.name):
                to_write['name'] = name_clean
            if to_write:
                existing.sudo().write(to_write)
            return existing

        # Documento oblig: dígitos
        digits = re.sub(r'\D', '', vat_raw)
        if not (name_clean and digits):
            raise UserError(_("Contacto requiere NOMBRE y DOCUMENTO (recibido: nombre='%s', doc='%s').") % (name_clean, vat_raw))

        # Busca existente con TUS reglas

        partner = self._find_existing_partner_by_vat(vat_raw, ident_type_id, do_nit=do_nit).with_context(active_test=False)



        # Función para normalizar NIT a un formato estable (ej. #######-DV o CO#######-DV si quieres)
        def _normalize_vat(v_raw, ident_id):
            d = re.sub(r'\D', '', self._s(v_raw))
            """ if not d:
                return self._s(v_raw)
            # → Para NIT: guarda con prefijo 'CO' y DV calculado: CO#########-D
            if do_nit and self._is_nit_type(ident_id) and len(d) >= 9:
                base = d[:9]
                dv = self._co_nit_check_digit(base)
                _logger.info("base %s __ %s",base, dv)
                return f"{base}"
            # Otros documentos: solo dígitos """
            # _logger.info(" DDDD %s",d)
            return d

        vat_norm = _normalize_vat(vat_raw, ident_type_id)

        base_vals = {
            'name': name_clean,
            'vat': vat_norm,
        }
        if ident_type_id:
            base_vals['l10n_latam_identification_type_id'] = ident_type_id

        if partner:
            # completa solo vacíos o pisa si overwrite=True
            to_write = {}
            for k, v in base_vals.items():
                if v in (None, '', False):
                    continue
                if overwrite or not getattr(partner, k, False):
                    if k == 'vat' and getattr(partner, 'vat', False):
                        old_d = re.sub(r'\D', '', self._s(partner.vat))
                        new_d = re.sub(r'\D', '', self._s(v))
                        if not overwrite and old_d and old_d != new_d:
                            continue  # no cambiar vat si difiere en dígitos
                    to_write[k] = v
            if to_write:
                partner.write(to_write)
            return partner

        # Crear si no existe — evitar validación VAT en import masivo
        try:
            # usamos el contexto para bypass del validador de VAT y sudo() para permisos
            partner = self.env['res.partner'].with_context(no_vat_validation=False).sudo().create(base_vals)
            return partner
        except Exception as e:
            _logger.warning("Creación de partner con bypass de VAT falló (%s). Intentando creación sin bypass: %s", e, base_vals)
            # Intento fallback: crear sin VAT si existe 'vat' en base_vals
            base_vals_fallback = base_vals.copy()
            base_vals_fallback.pop('vat', None)
            try:
                partner = self.env['res.partner'].sudo().create(base_vals_fallback)
                return partner
            except Exception:
                # último recurso: vuelve a levantar la excepción original para no esconder errores graves
                raise


    def _find_existing_partner_by_vat(self, raw_vat, ident_type_id=None, *, do_nit=False):
        """
        Busca partner por (tipo + dígitos del doc). Tolera formatos distintos.
        Respeta active_test=False para considerar archivados.
        Devuelve un res.partner (0 o 1).
        """
        # Partner = self.env['res.partner'].with_context(active_test=False).sudo()
        Partner = self.env['res.partner'].sudo()
        digits_only = re.sub(r'\D', '', self._s(raw_vat))
        if not digits_only:
            return Partner.browse()
        domain = [('vat', 'ilike', digits_only)]
        if ident_type_id:
            domain.insert(0, ('l10n_latam_identification_type_id', '=', ident_type_id))

        # candidatos por ilike y luego filtra por igualdad exacta de dígitos
        for p in Partner.search(domain, limit=100):
            pv = re.sub(r'\D', '', self._s(p.vat))
            if pv == digits_only:
                return p

        # último intento sin tipo (por si no vino o difiere)
        if ident_type_id:
            for p in Partner.search([('vat', 'ilike', digits_only)], limit=100):
                pv = re.sub(r'\D', '', self._s(p.vat))
                if pv == digits_only:
                    return p

        return Partner.browse()

    def _vat_variants(self, raw_vat, ident_type_id, do_nit):
        """Devuelve posibles variantes de VAT para evitar duplicados por formato."""
        d = re.sub(r'\D', '', self._s(raw_vat))
        cand = []
        # lo que viene tal cual
        if raw_vat:
            cand.append(self._s(raw_vat))
        # solo dígitos
        if d:
            cand.append(d)
        # variantes NIT si corresponde
        if do_nit and self._is_nit_type(ident_type_id) and len(d) == 9:
            dv = self._co_nit_check_digit(d)
            cand.extend([
                f"CO{d}-{dv}",  # formato final estándar
                f"{d}-{dv}",    # por si hay registros viejos sin 'CO'
                f"CO{d}",       # por si alguien guardó con 'CO' pero sin DV
            ])
        # quitar vacíos y duplicados preservando orden
        seen, res = set(), []
        for c in cand:
            if c and c not in seen:
                seen.add(c)
                res.append(c)
        return res
    def _co_nit_check_digit(self, nit_digits: str) -> int:
        """Calcula DV DIAN a partir de solo dígitos."""
        weights = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]

        nums = re.sub(r'\D', '', nit_digits)
        total = 0
        for i, d in enumerate(nums[::-1]):
            if i >= len(weights):
                break
            total += int(d) * weights[i]
        r = total % 11
        return r if r in (0, 1) else 11 - r

    def _is_nit_type(self, ident_type_id):
        if not ident_type_id:
            return False
        IL = self.env['l10n_latam.identification.type'].sudo().browse(ident_type_id)
        code = (getattr(IL, 'code', '') or getattr(IL, 'l10n_co_document_code', '') or '').upper()
        shortcut = (getattr(IL, 'shortcut', '') or getattr(IL, 'short_name', '') or '').upper()
        name = (IL.name or '').upper()
        return code == '31' or 'NIT' in (shortcut + name)


    def _update_partner_contact(
        self, partner, *, mobile=None, phone=None, email=None,
        street=None, street2=None, city=None, state_name=None, country_id=None
    ):
        """Escribe solo campos existentes. Sobrescribe si traemos dato."""
        if not partner:
            return
        P = partner.sudo()
        f = P._fields
        vals = {}
        if mobile and 'mobile' in f:
            vals['mobile'] = self._s(mobile)
        if phone and 'phone' in f:
            vals['phone'] = self._s(phone)
        if email and 'email' in f:
            vals['email'] = self._s(email)
        if street and 'street' in f:
            vals['street'] = self._s(street)
        if street2 and 'street2' in f:
            vals['street2'] = self._s(street2)
        if city and 'city' in f:
            vals['city'] = self._s(city)
        if country_id and 'country_id' in f:
            vals['country_id'] = country_id
        # state_id se resuelve por nombre/código si viene y hay país
        if state_name and 'state_id' in f and (country_id or P.country_id.id):
            sid = self._find_state(country_id or P.country_id.id, state_name)
            if sid:
                vals['state_id'] = sid
        if vals:
            P.write(vals)
    def _compose_street(self, street, urbanizacion, barrio):
        """
        Devuelve (street, street2) priorizando:
        - street: Dir_Correspondencia_Inq
        - street2: 'Urbanización X - Barrio Y' si existen
        """
        s1 = self._s(street)
        parts = []
        u = self._s(urbanizacion)
        b = self._s(barrio)
        if u:
            parts.append(f"Urbanización {u}")
        if b:
            parts.append(f"Barrio {b}")
        s2 = ' - '.join(parts) if parts else ''
        return s1, s2
    def _company_country_id(self):
        comp = self.env.user.company_id
        return comp.country_id.id if comp and comp.country_id else False

    def _get_latam_ident_type_id(self, raw):
        """
        Devuelve el ID de l10n_latam.identification.type según el valor del Excel.
        Soporta números ('31','13') y textos ('NIT','CC','Cédula', etc.), y
        se adapta a los campos disponibles en el modelo (code, shortcut, short_name,
        l10n_co_document_code, name).
        """
        IL = self.env['l10n_latam.identification.type'].sudo()
        txt = (self._s(raw) or '').strip()
        if not txt:
            return False

        upper = txt.upper()
        digits = ''.join(ch for ch in upper if ch.isdigit())  # p.ej. '31' de '31 ' o ' 31'
        country_id = self._company_country_id()

        # Dominio por país (o sin país), igual al que ves en la vista.
        dom_country = ['|', ('country_id', '=', False)]
        if country_id:
            dom_country += [('country_id', '=', country_id)]
        else:
            dom_country += [('country_id', '=', False)]

        # Mapeos comunes CO
        #   31 -> NIT
        #   13 -> Cédula de ciudadanía (CC)
        name_by_code = {
            '31': 'NIT',
            '13': 'Cédula de ciudadanía',  # usamos 'Cédula' genérico para ilike
        }
        shortcuts_by_code = {
            '31': ['NIT', 'RUT'],   # por si en tu DB usan 'NIT' o 'RUT' como shortcut
            '13': ['CC', 'CEDULA', 'CÉDULA'],
        }

        fields_avail = IL._fields

        def _try(dom):
            rec = IL.search(dom + dom_country, limit=1)
            return rec.id if rec else False

        # 1) Si tu modelo tiene 'code' y el excel trae dígitos (31/13), probar por code
        if digits:
            if 'code' in fields_avail:
                rid = _try([('code', '=', digits)])
                if rid:
                    return rid
            # Algunos localismos usan otro campo "código"
            for alt_code_field in ('l10n_co_document_code', 'internal_code'):
                if alt_code_field in fields_avail:
                    rid = _try([(alt_code_field, '=ilike', digits)])
                    if rid:
                        return rid

        # 2) Probar por shortcuts típicos si el modelo los tiene
        if 'shortcut' in fields_avail or 'short_name' in fields_avail:
            candidates = []
            if digits and digits in shortcuts_by_code:
                candidates += shortcuts_by_code[digits]
            # añade el texto crudo también
            candidates += [upper]
            for c in candidates:
                if 'shortcut' in fields_avail:
                    rid = _try([('shortcut', '=ilike', c)])
                    if rid:
                        return rid
                if 'short_name' in fields_avail:
                    rid = _try([('short_name', '=ilike', c)])
                    if rid:
                        return rid

        # 3) Probar por name (ilike), usando nombre mapeado si tenemos dígitos
        name_candidates = []
        if digits and digits in name_by_code:
            name_candidates.append(name_by_code[digits])
        name_candidates.append(txt)   # el texto tal cual del Excel

        for c in name_candidates:
            rid = _try([('name', 'ilike', c)])
            if rid:
                return rid

        # 4) Último intento: buscar cualquier coincidencia con el texto en varios campos
        for field in ('name', 'shortcut', 'short_name', 'l10n_co_document_code'):
            if field in fields_avail:
                rid = _try([(field, 'ilike', upper)])
                if rid:
                    return rid

        return False


    # --- NUEVO: buscar por dirección (solo usa campos existentes) ---
    def _property_by_address(self, addr, city=None):
        """Busca account.analytic.account por dirección y opcionalmente por municipio,
        usando SOLO campos que existan en el modelo. Si no hay campos de dirección, cae a name."""
        Analytic = self.env['account.analytic.account'].sudo()
        addr = self._s(addr)
        city = self._s(city)
        if not addr:
            return Analytic.browse()

        fields_avail = Analytic._fields

        # 1) Intentar por campos de dirección que puedan existir
        addr_fields = [f for f in ('x_property_geolocation', 'x_street', 'street') if f in fields_avail]
        city_fields = [f for f in ('city', 'x_city') if f in fields_avail and city]

        for faddr in addr_fields:
            domain = [(faddr, '=', addr)]
            if city_fields:
                # prueba con el primer campo de ciudad disponible
                domain = domain + [(city_fields[0], '=ilike', city)]
            rec = Analytic.search(domain, limit=1)
            if rec:
                return rec

        # 2) Si no hay campos de dirección o no hubo match, intenta por name
        rec = Analytic.search([('name', '=', addr)], limit=1)
        if rec:
            return rec
        return Analytic.search([('name', '=ilike', addr)], limit=1)


    def _parse_duration_months(self, v):
        s = self._s(v).lower()
        if not s:
            return False
        # 1) si hay dígitos, usa el primer número

        m = re.search(r'\d+', s)
        if m:
            try:
                return int(m.group(0))
            except Exception:
                pass
        # 2) mapa de palabras comunes
        mapa = {

            'uno': 1, 'una': 1, 'dos': 2, 'tres': 3, 'cuatro': 4, 'cinco': 5, 'seis': 6, 'siete': 7, 'ocho': 8, 'nueve': 9,
            'diez': 10, 'once': 11, 'doce': 12, 'quince': 15, 'dieciocho': 18,
            'veinte': 20, 'veintiun': 21, 'veintiuno': 21, 'veintidos': 22, 'veintitrés': 23, 'veintitres': 23, 'veinticuatro': 24,
            'treinta': 30, 'treinta y seis': 36, 'treintayseis': 36,
            'cuarenta y ocho': 48, 'cuarentayocho': 48,
            'sesenta': 60,
            # atajos típicos
            '1 mes': 1, '6 meses': 6, '12 meses': 12,

        }
        # normaliza espacios
        s = ' '.join(s.split())
        if s in mapa:
            return mapa[s]
        # intenta por primera palabra (p.e. "seis meses")
        first = s.split(' ')[0]
        return mapa.get(first, False)

    def _get_or_create_partner_by_vat(self, vat, name, ident_type_id=None, *, do_nit=False):
        Partner = self.env['res.partner'].sudo()
        name_clean = ' '.join(self._s(name).split())
        raw_vat = self._s(vat)
        # 🔒 Guardia: si ya existe por dígitos, reusar
        existing = self._search_partner_by_vat_digits(raw_vat)
        if existing:
            # Completar SOLO vacíos (sin cambiar VAT ni nombre si ya están)
            to_write = {}
            if ident_type_id and not existing.l10n_latam_identification_type_id:
                to_write['l10n_latam_identification_type_id'] = ident_type_id
            if name_clean and not existing.name:
                to_write['name'] = name_clean
            if to_write:
                existing.sudo().write(to_write)
            return existing
        
        # Documento debe tener dígitos
        digits_only = re.sub(r'\D', '', raw_vat)
        if not digits_only:
            raise UserError(_("Documento inválido para '%s' (no hay dígitos).") % name_clean)

        # Si piden NIT pero el tipo no es NIT, ignorar formateo NIT
        if do_nit and not self._is_nit_type(ident_type_id):
            _logger.warning("do_nit=True pero el tipo NO es NIT; se ignora. name=%s vat=%s", name_clean, raw_vat)
            do_nit = False

        if not raw_vat or not name_clean:
            raise UserError(_("Contacto requiere NOMBRE y DOCUMENTO (recibido: nombre='%s', doc='%s').") % (name_clean, raw_vat))
        # Normaliza a canónico solo si es NIT; si no, usa dígitos
        digits_only = re.sub(r'\D', '', self._s(vat))
        if not digits_only:
            raise UserError(_("Documento inválido para '%s' (no hay dígitos).") % name)

        partner = self._find_existing_partner_by_vat(vat, ident_type_id, do_nit=do_nit)

        # partner = self._find_existing_partner_by_vat(vat, ident_type_id)
        if partner:
            vals_to_write = {}

            # nombre
            name_clean = ' '.join(self._s(name).split())
            if partner.name != name_clean:
                vals_to_write['name'] = name_clean

            # tipo (si vino y difiere)
            current_type_id = partner.l10n_latam_identification_type_id.id if partner.l10n_latam_identification_type_id else False
            if ident_type_id and current_type_id != ident_type_id:
                vals_to_write['l10n_latam_identification_type_id'] = ident_type_id

            partner_digits = re.sub(r'\D', '', self._s(partner.vat))

            incoming_digits = re.sub(r'\D', '', self._s(raw_vat))

            if partner_digits == incoming_digits or not partner_digits:
                target_vat = incoming_digits
                if ident_type_id and self._is_nit_type(ident_type_id) and len(incoming_digits) >= 9:
                    base = incoming_digits[:9]
                    dv = self._co_nit_check_digit(base)
                    target_vat = f"{base}-{dv}"  # sin "CO"
                if self._s(partner.vat) != target_vat:
                    try:
                        partner.with_context(no_vat_validation=False).sudo().write({'vat': target_vat})
                    except Exception as e:
                        _logger.warning(
                            "No pude normalizar/asignar VAT para %s: %s (se deja igual).",
                            partner.display_name, e
        )

            return partner

        # --- Variantes para deduplicar (exactas) ---
        variants = self._vat_variants(raw_vat, ident_type_id, do_nit)
        if do_nit and self._is_nit_type(ident_type_id) and len(digits_only) != 9:
            variants.append(f"CO{digits_only}")
        variants = list(dict.fromkeys([v for v in variants if v]))

        # Buscar por cualquier variante exacta
        partner = Partner.search([('vat', 'in', variants)], limit=1)

        # --- Normalización a guardar ---
        vat_to_save = digits_only  # por defecto: solo dígitos
        base = digits_only[:9]
        dv = None
        if do_nit and self._is_nit_type(ident_type_id):
            if len(base) == 9:
                dv = self._co_nit_check_digit(base)
                vat_to_save = f"CO{base}-{dv}"
            else:
                _logger.warning("Se esperaba un NIT de 9 dígitos para '%s' (%s). Se guardan solo dígitos.", name_clean, raw_vat)

        # --- Fallback anti-duplicados SOLO si no encontró exacto ---
        # Captura registros guardados como "CO###.###.###-DV", "###-DV", "CO#########", etc.
        # --- Fallback anti-duplicados SOLO si no encontró exacto ---
        if not partner:
            acceptable_sane = {digits_only}
            if len(base) == 9 and dv is not None:
                acceptable_sane.add(base)
                if dv is None:
                    try:
                        dv = self._co_nit_check_digit(base)
                    except Exception:
                        dv = None

            # Construye patrones únicos para buscar por texto (evita OR mal formado)
            patterns = {digits_only}
            if len(base) == 9:
                patterns.add(base)
                if dv is not None:
                    dotted = f"{base[:3]}.{base[3:6]}.{base[6:9]}"
                    patterns.add(f"{base}-{dv}")
                    patterns.add(f"{dotted}-{dv}")

            # Crea cláusulas únicas
            clauses = [('vat', 'ilike', p) for p in patterns if p]
            # Búsqueda segura con OR prefijo
            if not clauses:
                candidates = self.env['res.partner'].sudo().browse()
            elif len(clauses) == 1:
                candidates = Partner.search([clauses[0]], limit=80)
            else:
                wide_domain = (['|'] * (len(clauses) - 1)) + clauses
                candidates = Partner.search(wide_domain, limit=80)

            for p in candidates:
                sane = re.sub(r'\D', '', self._s(p.vat))
                if sane in acceptable_sane:
                    partner = p
                    break
        # --- Crear ---
        vals = {'name': name_clean, 'vat': vat_to_save}
        if ident_type_id:
            vals['l10n_latam_identification_type_id'] = ident_type_id


        try:
            return self.env['res.partner'].with_context(no_vat_validation=False).sudo().create(vals)
        except Exception as e:
            _logger.warning("Create partner con VAT %s falló: %s. Intentando sin VAT.", vals.get('vat'), e)

    def _upsert_clause_var(self, so_id, key, value):
        """Crea/actualiza cláusula del contrato (modelo clause.var: contract_id/key/value)."""
        if value is None or (isinstance(value, str) and self._s(value) == ''):
            return
        ClauseVar = self.env['clause.var'].sudo()
        clause = ClauseVar.search([('contract_id', '=', so_id), ('key', '=', key)], limit=1)
        vals = {'value': self._s(value)}
        if clause:
            clause.write(vals)
        else:
            ClauseVar.create({'contract_id': so_id, 'key': key, **vals})


    def _set_property_location(
        self, analytic, addr, city, urbanizacion=None, barrio=None, *, tipo_prop=None, sector=None, overwrite=True
    ):
        """
        Fragmenta una dirección CO y escribe campos x_* si existen.
        + x_property_type (selection) desde 'tipo_prop'
        + Sector/Barrio desde 'sector' en el primer campo disponible

        """
        if not analytic:
            return

        A = analytic.sudo()
        f = A._fields

        def _s(v): return (v or "").strip()

        addr = _s(addr)
        city = _s(city)
        urb = _s(urbanizacion)
        bar = _s(barrio)

        # helpers
        def _empty_or_overwrite(field):
            return (field in f) and (overwrite or not getattr(A, field, False))

        def _set(field, value, vals):
            if _empty_or_overwrite(field) and value not in (None, ''):
                vals[field] = value


        # --- normalizadores y tablas ---
        CANON = {
            # Calle
            'CL': 'CL', 'CLL': 'CL', 'CALLE': 'CL',
            # Carrera
            'CR': 'CR', 'CRA': 'CR', 'CARRERA': 'CR',
            # Avenida
            'AV': 'AV', 'AVENIDA': 'AV',
            # Diagonal
            'DG': 'DG', 'DIAGONAL': 'DG',
            # Transversal
            'TV': 'TV', 'TR': 'TV', 'TRANSVERSAL': 'TV',
            # Otros frecuentes
            'CIR': 'CIR', 'CIRCULAR': 'CIR',
            'CT': 'CT', 'CARRETERA': 'CT',
            'AUT': 'AUT', 'AUTOPISTA': 'AUT',
        }
        ABBR2FAMILY = {
            'CL': 'calle', 'CR': 'carrera', 'AV': 'avenida',
            'DG': 'diagonal', 'TV': 'transversal',
            'CIR': 'circular', 'CT': 'carretera', 'AUT': 'autopista',
        }
        FAMILY_ALIASES = {
            'calle': {'cl', 'cll', 'calle'},
            'carrera': {'cr', 'cra', 'carrera'},
            'avenida': {'av', 'avenida'},
            'diagonal': {'dg', 'diagonal'},
            'transversal': {'tv', 'tr', 'transversal'},
            'circular': {'cir', 'circular'},
            'carretera': {'ct', 'carretera'},
            'autopista': {'aut', 'autopista'},
        }

        def _tipo_via2(abbr):
            if abbr == 'CL': return 'CR'
            if abbr == 'CR': return 'CL'
            if abbr == 'AV': return 'CL'
            if abbr == 'DG': return 'CR'
            if abbr in ('TV', 'TR'): return 'CL'
            return ''

        def _normalize_txt(s):
            t = _s(s)
            t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
            return re.sub(r'\s+', ' ', t).strip().lower()


        def _get_selection_list(field_name):
            if field_name not in f:
                return []
            sel_def = A._fields[field_name].selection
            if callable(sel_def):
                try:
                    return list(sel_def(A))
                except TypeError:
                    return list(sel_def())
            return list(sel_def or [])
        
        def _selection_key(field_name, *candidates):
            sel = _get_selection_list(field_name)
            if not sel:
                return None
            labels = {_normalize_txt(lbl): key for key, lbl in sel}
            keys = {_normalize_txt(str(key)): key for key, _ in sel}
            for raw in candidates:
                n = _normalize_txt(raw)
                if not n:
                    continue
                if n in keys:
                    return keys[n]
                if n in labels:
                    return labels[n]
                # contains
                for lbln, key in labels.items():
                    if n == lbln or n in lbln or lbln in n:
                        return key
            return None
        
        # -------------------------------
        # 1) Parseo (robusto si addr vacío)
        # -------------------------------
        tv_abbr = ''        # 'CL','CR','AV','...' (canónica)
        nombre_via = ''
        vs = ''             # número secundario
        num = ''            # número principal
        comp = ''           # complemento
        tokens = []

        if addr:
            s_upper = addr.upper()
            tokens = s_upper.split()

            # a) Primer token -> abreviatura canónica
            first = (tokens[0] if tokens else '').upper()
            tv_abbr = CANON.get(first, '')  # '' si no reconoce

            # b) nombre de vía si hay segundo token y se reconoció tipo
            tail_tokens = tokens[1:]
            if tv_abbr and len(tokens) >= 2:
                nombre_via = tokens[1]
                tail_tokens = tokens[2:]

            # c) heurística simple: siguientes dos tokens alfanum como vs y num
            alnums = [t for t in tail_tokens if re.match(r'^[0-9A-Z]+$', t)]

        


        # ---- tablas de vía ----
        CANON = {
            'CL': 'CL', 'CLL': 'CL', 'CALLE': 'CL',
            'CR': 'CR', 'CRA': 'CR', 'CARRERA': 'CR',
            'AV': 'AV', 'AVENIDA': 'AV',
            'DG': 'DG', 'DIAGONAL': 'DG',
            'TV': 'TV', 'TR': 'TV', 'TRANSVERSAL': 'TV',
            'CIR': 'CIR', 'CIRCULAR': 'CIR',
            'CT': 'CT', 'CARRETERA': 'CT',
            'AUT': 'AUT', 'AUTOPISTA': 'AUT'
        }

        def _tipo_via2(abbr):
            return {'CL': 'CR', 'CR': 'CL', 'AV': 'CL', 'DG': 'CR', 'TV': 'CL', 'TR': 'CL'}.get(abbr, '')

        # ------------------ 1) Parseo dirección ------------------
        tv_abbr = nombre_via = vs = num = comp = ''
        if addr:
            toks = addr.upper().split()
            first = (toks[0] if toks else '').upper()
            tv_abbr = CANON.get(first, '')
            tail = toks[1:]
            if tv_abbr and len(toks) >= 2:
                nombre_via = toks[1]
                tail = toks[2:]
            alnums = [t for t in tail if re.match(r'^[0-9A-Z]+$', t)]

            if len(alnums) >= 1:
                vs = alnums[0]
            if len(alnums) >= 2:
                num = alnums[1]
            used = {x for x in (vs, num) if x}
            comp_parts = [t for t in tail if t not in used]
            comp = " ".join(comp_parts).strip()
            extras = []
            if urb:
                extras.append(f"Urbanización {urb}")
            if bar:
                extras.append(f"Barrio {bar}")
            if extras:
                comp = (f"{comp} - " if comp else "") + " - ".join(extras)

        tv2_abbr = _tipo_via2(tv_abbr)

        # mapa abreviatura -> familia
        ABBR2FAMILY = {
            'CL': 'calle', 'CR': 'carrera', 'AV': 'avenida',
            'DG': 'diagonal', 'TV': 'transversal', 'TR': 'transversal',
            'CIR': 'circular', 'CT': 'carretera', 'AUT': 'autopista',
        }

        fam1 = ABBR2FAMILY.get(tv_abbr, '')
        fam2 = ABBR2FAMILY.get(tv2_abbr, '')

        # ------------------ 2) Escribir valores ------------------
        vals = {}
        # Pasa abreviatura y familia como candidatos (clave o etiqueta)
        # _set('x_tipo_via', _selection_key('x_tipo_via', tv_abbr, fam1), vals)
        # _set('x_tipo_via2', _selection_key('x_tipo_via2', tv2_abbr, fam2), vals)
        # Resuelve clave para tipo_via (tv_abbr → CL/CR/AV/DG/TV/TR)
        # Resuelve key real de x_tipo_via
        # --- Resolver x_tipo_via y derivar x_tipo_via2 (opuesto robusto) ---

        # 1) Resuelve key real de x_tipo_via con varias pistas
        # ------------------ Tipos de vía: simple y robusto ------------------
        

        # ------------------ Tipos de vía: robusto y mínimo ------------------
        # --- Via principal y secundaria (forzado por mapeo explícito) ---
        vals = vals or {}

        # Abreviatura principal ya calculada en tu código:
        abbr = (tv_abbr or '').strip().upper()
        if not abbr:
            raw = (getattr(A, 'x_street_name', '') or '').split()[:1] or ['']
            raw = (raw[0] or '').strip().upper()
            if raw in ('CALLE','CL','CLL'): abbr = 'CL'
            elif raw in ('CARRERA','CR','CRA'): abbr = 'CR'
            elif raw in ('AVENIDA','AV'): abbr = 'AV'
            elif raw in ('DIAGONAL','DG'): abbr = 'DG'
            elif raw in ('TRANSVERSAL','TV','TR'): abbr = 'TV'

        # Mapa de la abreviatura canónica -> key real del selection
        KEYMAP = {'CL': 'cll', 'CR': 'cra', 'AV': 'av', 'DG': 'dg', 'TV': 'tv', 'TR': 'tv'}

        def _set_sel(field, ab):
            key = KEYMAP.get(ab)
            if key and field in f and (overwrite or not getattr(A, field, False)):
                vals[field] = key

        # Principal
        _set_sel('x_tipo_via', abbr)

        # Secundaria = opuesto básico (ya tienes _tipo_via2)
        opp = _tipo_via2(abbr) or ''
        _set_sel('x_tipo_via_2', opp)

        # _logger.info("LOGGER via1=%s(%s) via2=%s(%s)",
        #             abbr, KEYMAP.get(abbr), opp, KEYMAP.get(opp))


        _set('x_nombre_via', nombre_via, vals)
        _set('x_numero_secundario', vs, vals)
        _set('x_numero_principal', num, vals)
        _set('x_complemento', comp, vals)


        if (tv_abbr or nombre_via):
            _set('x_street_name', f"{tv_abbr} {nombre_via}".strip(), vals)
        if (vs or num):
            _set('x_street_number', f"{vs}-{num}" if (vs or num) else '', vals)


        for candidate in ('x_property_geolocation', 'x_street', 'street'):
            if _empty_or_overwrite(candidate) and addr:
                vals.setdefault(candidate, addr)
                break

        for candidate in ('city', 'x_city'):
            if _empty_or_overwrite(candidate) and city:
                vals[candidate] = city
                break


        # Fallback nombre
        # ------------------ 3) Tipo de propiedad + Sector ------------------
        # Tipo (selection x_property_type) a partir de 'tipo_prop' (Excel “Clase”)
        if 'x_property_type' in f and _empty_or_overwrite('x_property_type'):
            # alias rápidos
            ALIASES = {
                'Apartamento': {'apartamento', 'apto'},
                'Apartaestudio': {'apartaestudio', 'aparta estudio'},
                'Local': {'local'},
                'Bodega': {'bodega'},
                'Oficina/Consultorio': {'oficina', 'consultorio', 'oficina-consultorio', 'oficina/consultorio'},
                'Casa': {'casa'},
                'Estudio': {'estudio'},
                'Espacio comercial/Publicidad': {'espacio comercial', 'publicidad', 'espacio comercial/publicidad'},
                'Terreno': {'terreno', 'lote', 'lote/terreno'},
                'Casa lote': {'casa lote'},
                'Garaje': {'garaje'},
                'Habitación': {'habitacion', 'habitación', 'hab'},
                'Parqueadero': {'parqueadero', 'parqueo', 'parq'},
                'Terraza': {'terraza'},
                'Finca': {'finca', 'casa finca', 'casa-finca'},
            }
            cands = [tipo_prop] if tipo_prop else []
            tnorm = _normalize_txt(tipo_prop)
            for label, al in ALIASES.items():
                if tnorm and tnorm in al:
                    cands.append(label)
            key = _selection_key('x_property_type', *cands)
            if key is not None:
                vals['x_property_type'] = key

        # Sector/Barrio desde 'sector' → usa el primer campo existente
        barrio_field = next((fn for fn in ('x_sector', 'sector', 'x_barrio_sector', 'x_barrio', 'x_neighborhood') if fn in f), None)
        if barrio_field and _empty_or_overwrite(barrio_field) and _s(sector):
            vals[barrio_field] = _s(sector)
        # ------------------ 4) Fallback nombre ------------------

        if not vals and addr and ('name' in f) and (overwrite or not A.name):
            vals['name'] = addr

        if vals:
            A.write(vals)

    def _s(self, v):
        """String seguro (no revienta con int/None)."""
        if v is None:
            return ''
        # deja números como string sin .strip en ints/floats
        if isinstance(v, (int, float)):
            return str(v)
        return str(v).strip()

    def _get_ws(self):
        """Abrir workbook desde self.file y devolver la hoja pedida o la primera."""
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise UserError(_("Falta la librería 'openpyxl' (pip install openpyxl).")) from e
        if not self.file:
            raise UserError(_("Adjunta un archivo .xlsx"))

        data = base64.b64decode(self.file)
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            return wb[self.sheet_name]
        return wb[wb.sheetnames[0]]

    @staticmethod
    def _norm_header(s):
        import unicodedata, re
        txt = str(s or '').strip()
        # sin tildes
        txt = ''.join(c for c in unicodedata.normalize('NFKD', txt) if not unicodedata.combining(c))
        txt = re.sub(r'\s+', ' ', txt).lower()
        return txt

    def _header_map(self, ws):
        """Devuelve (headers_list, headers_rev_dict) normalizados a partir de la fila 1."""
        headers = [self._norm_header(c.value) for c in ws[1]]
        headers_rev = {h: i for i, h in enumerate(headers)}
        return headers, headers_rev

    @staticmethod
    def _get(row, headers_rev, header_name, default=''):
        """Leer valor por nombre normalizado de encabezado."""
        idx = headers_rev.get(ContractExcelImportWizard._norm_header(header_name))
        if idx is None or idx >= len(row):
            return default
        return row[idx] if row[idx] is not None else default

    @staticmethod
    def _to_float(v, default=0.0):
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if not s:
            return default
        # tolerante: quitar espacios, miles y normalizar coma/punto
        s = s.replace(' ', '')
        # Heurística simple que funciona bien para “1.234,56” y “1,234.56”
        if ',' in s and '.' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        else:
            s = s.replace('.', '').replace(',', '.')
        try:
            return float(s)
        except Exception:
            return default

    @staticmethod
    def _parse_date(v):
        if not v:
            return False
        if isinstance(v, datetime):
            return v.date()
        if hasattr(v, 'date'):
            try:
                return v.date()
            except Exception:
                pass
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        return False

    def _scope_from_text(self, txt):
        t = (self._s(txt) or '').strip().lower()
        if not t:
            return 'rental'  # vacío -> Renta
        if t in ('arriendo', 'alquiler', 'arrendamiento', 'rental'):
            return 'rental'
        if t in ('owner', 'administrativo', 'administracion', 'administración'):
            return 'owner'
        # por defecto, si viene algo raro, considéralo rental
        return 'rental'

    def _ensure_sale_order(self, name):
        return self.env['sale.order'].sudo().search([('name', '=', name)], limit=1)

    # -------------------------------------------------------
    # Acción principal (simple, robusta y sin helpers muertos)
    # -------------------------------------------------------
    
    def _default_rent_charge_line(self, canon=0):
        """Si no hay líneas, usa canon_property con el producto 'Tarifa de alquiler'."""
        canon = canon if canon>0 else 0
        if canon <= 0:
            return []
        try:
            product = self.env['sale.order']._get_product_tarifa_alquiler()
            return [{'product_id': product.id, 'amount': canon}]
        except Exception as e:
            _logger.exception("ECOERP DEBUG: fallo resolviendo producto Tarifa de alquiler")
            # Si quieres abortar aquí:
            # raise
            # O registrar y continuar:
            return []
        
    def _default_admin_charge_line(self, administracion_ph=0, monto_administracion=0):
        administracion_ph = administracion_ph if administracion_ph>0 else 0
        monto_administracion = monto_administracion if monto_administracion>0 else 0
        if administracion_ph <= 0 or monto_administracion <= 0:
            return []
        try:
            product = self.env['sale.order']._get_product_administracion_ph()
            # value = (administracion_ph*monto_administracion)/100
            value = monto_administracion
            return [{'product_id': product.id, 'amount': value}]
        except Exception as e:
            _logger.exception("ECOERP DEBUG: fallo resolviendo producto Tarifa de alquiler")
            # Si quieres abortar aquí:
            # raise
            # O registrar y continuar:
            return []
        
    def _check_owner_percentages(self, lines_log):
        """Valida la suma de participación de propietarios por contrato."""
        if lines_log is None:
            lines_log = []
        query = """
            SELECT
                so.id                  AS so_id,
                so.name                AS contract_number,
                aaa.name               AS property_name,
                COUNT(aaol.id)         AS owners_count,
                COALESCE(SUM(
                    COALESCE(aaol.participation_percent,
                            aaol.beneficial_porcentage, 0)
                ), 0)                  AS pct_total
            FROM sale_order so
            LEFT JOIN account_analytic_account aaa
                ON aaa.id = so.x_account_analytic_account_id
            LEFT JOIN account_analytic_account_owner_line aaol
                ON aaol.analytic_account_id = aaa.id
            WHERE aaa.id IS NOT NULL
            GROUP BY so.id, so.name, aaa.name
            HAVING ABS(COALESCE(SUM(
                    COALESCE(aaol.participation_percent,
                            aaol.beneficial_porcentage, 0)
                ), 0) - 100) > 0.01
            ORDER BY so.name;
        """
        self.env.cr.execute(query)
        results = self.env.cr.dictfetchall()
        header = "VALIDACIÓN DE REGISTRO DE PROPIETARIOS"
        # _logger.info(header)
        lines_log.append(header)
        if not results:
            msg = "✅ Todas las propiedades tienen 100% de participación total."
            # _logger.info(msg)
            lines_log.append(msg)
            return
        warn = "⚠️ Propiedades con participación diferente a 100%:"
        _logger.warning(warn)
        lines_log.append(warn)
        for rec in results:
            missing = round(100 - rec['pct_total'], 2)
            # Opción A: logging con placeholders (escapando % como %%):
            _logger.warning(
                "Contrato: %s | Propiedad: %s | %s propietarios | Total: %.2f%% | Diferencia: %.2f%%",
                rec['contract_number'], rec['property_name'], rec['owners_count'],
                rec['pct_total'], missing
            )

            # Opción B (para lines_log): usa f-string (aquí NO se escapa el %)
            line = (f"Contrato: {rec['contract_number']} | Propiedad: {rec['property_name']} | "
                    f"{rec['owners_count']} propietarios | Total: {float(rec['pct_total']):.2f}% | "
                    f"Diferencia: {missing:.2f}%")
            lines_log.append(line)
        return lines_log
    
    def action_import(self):
        self.ensure_one()

        created = updated = skipped = simulated = 0
        lines_log = []

        try:
            ws = self._get_ws()
            headers, headers_rev = self._header_map(ws)

            # alias mínimos esperados (usa los nombres de tu Excel grande)
            C = {
                'contrato': 'Contrato',
                'tipo': 'TipoContrato',
                'prop_code': 'Propiedad / Código',
                'dir': 'Dirección del inmueble',
                'mun': 'Municipio del inmueble',
                'canon': 'Canon',
                'canon_let': 'canon_en_letras',
                'comision': 'Porcentaje_comision',
                'arr_nom': 'Nombre_Inquilino',
                # 👇 usa tu header nuevo y largo
                'arr_vat': 'Nro_id_Inquilino',
                'ini': 'fecha_contrato',
                'fin': 'Fecha_fin_contrato',
                'dur': 'Duracion_Contrato',
                'porcent_prop': 'Participacion_Prop',
                'porcent_prop2': 'Participacion_Prop2',
                'l10n_latam_identification_type_id':'l10n_latam_identification_type_id',
                'comision_inmob':'comision_inmob',
                'Porcentaje_administracion':'Porcentaje_administracion',
                'Valor_admin':'Valor_admin',
                'Fecha_Incremento':'Fecha_Incremento'
            }


            # Fallbacks de encabezados comunes
            fallback = {
                'Propiedad / Código': ['Cons_Inm', 'Codigo_Interno', 'Propiedad / Codigo', 'Propiedad / código'],
                'Dirección del inmueble': ['Dir_Inmueble', 'Direccion del inmueble', 'Direccion Inmueble'],
                'Municipio del inmueble': ['Municipio_Inmueble', 'Municipio del Inmueble'],
                'canon_en_letras': ['Canon en letras', 'Canon (letras)'],
                'Porcentaje_comision': ['Comision', '% Comision', 'Porcentaje comision', 'Comisión'],
                'Nombre_Inquilino': ['Arrendatario (nombre)', 'Inquilino', 'Arrendatario'],
                'Nro_id_Inquilino': ['Nro_id_Inquilino', 'Nro_id_Inquilino (obligatorio para vincular/crear)', 'Documento arrendatario', 'CC Inquilino', 'NIT Inquilino'],
                'fecha_contrato': ['Fecha inicio', 'Fecha Inicio Contrato'],
                'Fecha_fin_contrato': ['Fecha fin', 'Fecha Fin Contrato'],
                'Duracion_Contrato': ['Vigencia (meses)', 'Vigencia'],
                'Canon':['Canon', 'Arriendo'],
                'Participacion_Prop': ['Participacion','Porcentaje (%)', 'Beneficiario', 'titularidad', 'Beneficio'],
                'Participacion_Prop2': ['Participacion2', 'Participacion_2', 'Porcentaje2 (%)', 'Porcentaje_2 (%)',  'Beneficiario', 'titularidad2', 'titularidad_2', 'Beneficio'],
                'l10n_latam_identification_type_id': ['Tipo_Dcto_Inquilino'],
                'comision_inmob':['Porcentaje_comision', 'comision_inmobiliaria (%)'],
                'Porcentaje_administracion':['administracion (%)','Porcentaje_administracion'],
                'Valor_admin':['Comision_Admon','administracion','ValorAdministracion'],
                'Fecha_Incremento':['Fecha_Incremento','incremento']
            }

            # resolver alias en headers_rev
            def has(h):
                return self._norm_header(h) in headers_rev

            for target, aliases in fallback.items():
                if not has(target):
                    for a in aliases:
                        if has(a):
                            headers_rev[self._norm_header(target)] = headers_rev[self._norm_header(a)]
                            break

            # Dentro del loop de filas del Excel (antes de crear/escribir):
            seen_cons = set()   # decláralo antes del for, una vez por import
            errors = []         # acumula errores para reportar al final

            # loop de filas (valores crudos, puede venir datetime/float/str)
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # ... dentro del for i, row in enumerate(...):
                so_name = self._s(self._get(row, headers_rev, C['contrato']))
                if not so_name:
                    skipped += 1
                    lines_log.append(f"Fila {i}: sin número de contrato → omitido.")
                    continue

                tipo_raw = self._s(self._get(row, headers_rev, C['tipo']))
                scope = self._scope_from_text(tipo_raw or 'rental')

                dir_inm = self._s(self._get(row, headers_rev, C['dir']))
                mun_inm = self._s(self._get(row, headers_rev, C['mun']))
                urb_inm = self._s(self._get(row, headers_rev, 'Urbanizacion_Inmueble'))
                
                # por fila:
                cons = self._s(self._get(row, headers_rev, 'Cons_Inm'))

                canon_num = self._to_float(self._get(row, headers_rev, C['canon']))  # numérico
                canon_let = self._s(self._get(row, headers_rev, C['canon_let']))
                comision_txt = self._s(self._get(row, headers_rev, C['comision']))

                arr_nom = self._s(self._get(row, headers_rev, C['arr_nom']))
                arr_vat = self._s(self._get(row, headers_rev, C['arr_vat']))

                f_ini = self._parse_date(self._get(row, headers_rev, C['ini']))
                f_fin = self._parse_date(self._get(row, headers_rev, C['fin']))
                # vig_meses = self._s(self._get(row, headers_rev, C['dur']))
                vig_meses_txt = self._s(self._get(row, headers_rev, C['dur']))
                vig_meses = self._parse_duration_months(vig_meses_txt)
                porcent_prop = self._to_float(self._get(row, headers_rev, C['porcent_prop']))
                comision_inmob = self._to_float(self._get(row, headers_rev, C['comision_inmob']))
                administracion_ph = self._to_float(self._get(row, headers_rev, C['Porcentaje_administracion']))
                monto_administracion = self._to_float(self._get(row, headers_rev, C['Valor_admin']))      
                date_increment = self._parse_date(self._get(row, headers_rev, C['Fecha_Incremento']))                 
                # _logger.info("\n COMPUTEEEEE: %s", date_increment)
                # print("\n COMPUTEEEEE: %s", date_increment)

                # ===== ENTIDADES =====
                # Mantengo tu arrendatario y propiedad TAL CUAL
                tenant = None
                owner = None
                beneficiary = None
                prop = None

                if not self.simulate:
                    # ARRRENDATARIO
                    tenant = None
                    if arr_nom and arr_vat:
                        tenant_doc_type_raw = self._s(self._get(row, headers_rev, 'l10n_latam_identification_type_id')) # por obligación debe ser el de la dian
                        tenant_ident_id = self._get_latam_ident_type_id(tenant_doc_type_raw)
                        tenant_do_nit = bool(tenant_ident_id and self._is_nit_type(tenant_ident_id))

                        tenant = self._upsert_partner_by_vat(
                            arr_nom, arr_vat,
                            ident_type_id=tenant_ident_id,
                            do_nit=tenant_do_nit,
                            overwrite=False,  # no pisar datos ya existentes
                        )

                    # ===== DATOS DE CONTACTO DEL INQUILINO =====
                    t_mobile = self._s(self._get(row, headers_rev, 'Celular_Inq'))
                    t_email = self._s(self._get(row, headers_rev, 'Email_Inq'))

                    # Dirección de correspondencia del inquilino
                    t_dir = self._s(self._get(row, headers_rev, 'Dir_Correspondencia_Inq'))
                    t_city = self._s(self._get(row, headers_rev, 'Municipio_Correspondencia_Inq'))
                    t_urb = self._s(self._get(row, headers_rev, 'Urbanizacion_Correspondencia_Inq'))
                    t_barr = self._s(self._get(row, headers_rev, 'Barrio_Correspondencia_Inq'))

                    # Componer street/street2
                    street, street2 = self._compose_street(t_dir, t_urb, t_barr)

                    # País por defecto: el de la compañía
                    company_country_id = self._company_country_id()

                    if tenant:
                        self._update_partner_contact(
                            tenant,
                            mobile=t_mobile,
                            email=t_email,
                            street=street,
                            street2=street2,
                            city=t_city,
                            # state_name=t_state or False,
                            country_id=company_country_id
                        )
                        
                    
                    # antes de crear la propiedad
                    Plan = self.env['account.analytic.plan'].sudo()
                    plan = Plan.browse(2)
                    if not plan.exists():
                        # sé flexible con el nombre (a veces es "Properties")
                        plan = Plan.search([('name', 'ilike', 'Properties')], limit=1)
                    plan_id = plan.id or False
                    prop = False
                    
                    
                    # 1) Si hay consecutivo, úsalo como clave primaria
                    if cons:
                        prop = self.env['account.analytic.account'].sudo().search([
                            ('x_is_property', '=', True),
                            ('x_cons_inm', '=', cons),
                        ], limit=1)

                    # 2.1) Chequeo intra-archivo
                    if cons:
                        if cons in seen_cons:
                            errors.append(f"Duplicado en el archivo: Cons_Inm {cons}")
                            # salta esta fila
                            continue
                        seen_cons.add(cons)
                        
                    # 2.3) Si no hay consecutivo, cae a la heurística por dirección
                    if not prop:
                        prop = self._property_by_address(dir_inm, mun_inm)

                    # 2.4) Crear si no existe
                    if not prop and dir_inm:
                        # OJO: si cons viene, se guarda y el SQL constraint garantiza unicidad                        
                        plan = self.env.ref('industry_real_estate.analytic_plan_properties', raise_if_not_found=False) \
                            or self.env['account.analytic.plan'].search([('name', '=', 'Properties')], limit=1)
                        prop = self.env['account.analytic.account'].sudo().create({
                            'name': dir_inm,
                            'x_is_property': True,
                            'x_cons_inm': cons or False,
                            'plan_id': plan.id or False,
                            'canon':canon_num,
                        })

                    # 2.5) Si existe y vino Cons_Inm:
                    if prop and cons:
                        # Si el prop ya tiene un consecutivo distinto → ERROR duro (no tocamos nada)
                        if prop.x_cons_inm and prop.x_cons_inm != cons:
                            errors.append(
                                f"Conflicto: la dirección coincide con una propiedad con Cons_Inm={prop.x_cons_inm}, "
                                f"pero el Excel trae {cons}."
                            )
                            continue
                        # Si no tenía, asígnalo (una vez). Si ya lo tiene igual, no hacemos nada.
                        if not prop.x_cons_inm:
                            prop.sudo().write({'x_cons_inm': cons})

                    if prop and urb_inm and dir_inm:
                        building = self._ensure_building_by_name_address(urb_inm, dir_inm)
                        # enlazar en la propiedad si existe el campo Many2one
                        if building and 'x_building_id' in prop._fields and not prop.x_building_id:
                            prop.sudo().write({'x_building_id': building.id})

                    if prop:
                        self._set_property_location(
                            prop, dir_inm, mun_inm,
                            urbanizacion=self._s(self._get(row, headers_rev, 'Urbanizacion_Inmueble')),
                            barrio=self._s(self._get(row, headers_rev, 'Barrio_inmueble')),
                            tipo_prop=self._s(self._get(row, headers_rev, 'Clase')),
                            sector=self._s(self._get(row, headers_rev, 'Barrio_inmueble'))
                        )  # <<--- AQUI clave para actualizar sector y tipo de propiedad
                        
                    # justo después de resolver `prop`:
                    blocks = self._iter_owner_blocks(row, headers_rev)

                    # fila de la segunda plantilla (sin canon ni inquilino)
                    if self._is_owners_only_row(row, headers_rev, C):
                        self._owners_only_branch(i=i, prop=prop, blocks=blocks, lines_log=lines_log, so_name=so_name)
                        continue  # no intentes crear sale.order
                    
                    # ======== FISCALIZACION =======
                    def _norm(s):
                        s = self._s(s or '')
                        return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c)).strip().lower()

                    def _col(base_label, i):
                        """Primero intenta con sufijo ' {i}', si no existe cae a la etiqueta base (retrocompatible)."""
                        val = self._get(row, headers_rev, f'{base_label} {i}')
                        if val in (None, ''):
                            val = self._get(row, headers_rev, base_label)
                        return self._s(val)

                    Bank = self.env['res.bank'].sudo()
                    PBank = self.env['res.partner.bank'].sudo()

                    for i in range(1, 6):
                        # --- Identidad del propietario i ---
                        owner_name = _col('Nombre Propietario', i)
                        owner_vat  = _col('Cédula/Nit Propietario', i)
                        owner_doc  = _col('Tipo doc prop', i)

                        if not (owner_name and owner_vat):
                            continue  # nada que procesar para este índice

                        owner_ident_id = self._get_latam_ident_type_id(owner_doc)
                        owner_do_nit   = bool(owner_ident_id and self._is_nit_type(owner_ident_id))

                        # Upsert del partner (por NIT)
                        owner = self._upsert_partner_by_vat(
                            owner_name, owner_vat,
                            ident_type_id=owner_ident_id, do_nit=owner_do_nit, overwrite=False,
                        )
                        if not owner:
                            continue

                        # ---------------- Datos bancarios del propietario i ----------------
                        cta      = _col('Consignar_cta_Prop', i)
                        ach_code = _col('Codigo_ACH_Prop', i)      # BIC/ACH si aplica
                        tipo_cta = _col('Tipo_Cuenta_Prop', i)     # "Ahorros"/"Corriente" (si tienes x_tipo_cuenta)
                        bank_name = _col('Banco_Prop', i)

                        if cta:
                            # 1) Resolver/crear banco
                            bank = Bank.search([('bic', '=', ach_code)], limit=1) if ach_code else Bank.browse()
                            if not bank and bank_name:
                                bank = Bank.search([('name', '=ilike', bank_name)], limit=1)
                            if not bank and (bank_name or ach_code):
                                bank_vals = {'name': bank_name or (ach_code or 'Banco')}
                                if ach_code:
                                    bank_vals['bic'] = ach_code
                                bank = Bank.create(bank_vals)

                            # 2) Upsert de cuenta bancaria por (partner, acc_number)
                            dom = [('partner_id', '=', owner.id), ('acc_number', '=', cta)]
                            pb = PBank.search(dom, limit=1)

                            vals_pb = {
                                'partner_id': owner.id,
                                'acc_number': cta,
                                'currency_id': self.env.company.currency_id.id,  # requerido por vista
                                'acc_holder_name': owner.name,
                            }
                            if bank:
                                vals_pb['bank_id'] = bank.id
                            # Si tienes un campo personalizado para tipo:
                            if 'x_tipo_cuenta' in PBank._fields and tipo_cta:
                                vals_pb['x_tipo_cuenta'] = tipo_cta

                            if pb:
                                update_pb = {}
                                for k, v in vals_pb.items():
                                    if not getattr(pb, k, False) and v:
                                        update_pb[k] = v
                                if update_pb:
                                    pb.write(update_pb)
                            else:
                                PBank.create(vals_pb)

                        # ---------------- Régimen y Posición fiscal del propietario i ----------------
                        vals_owner = {}

                        # Régimen (l10n_co_edi_fiscal_regimen)
                        reg_raw = _col('Regimen_Propietario', i)
                        t = _norm(reg_raw)
                        regimen_code = '49'
                        if t:
                            # acepta variantes: "régimen común", "comun", etc.
                            if 'comun' in t:
                                regimen_code = '48'   # Responsable de IVA
                            elif 'simplificado' in t:        # "simplificado", "simple", etc.
                                regimen_code = '49'   # No responsable (No Aplica)

                        if regimen_code and 'l10n_co_edi_fiscal_regimen' in owner._fields and owner.l10n_co_edi_fiscal_regimen != regimen_code:
                            vals_owner['l10n_co_edi_fiscal_regimen'] = regimen_code

                        # Posición fiscal (Many2one estándar)
                        reg_raw_pos = _col('Posicion_fiscal', i)  # si tu Excel lo trae por cada propietario
                        if reg_raw_pos and 'property_account_position_id' in owner._fields:
                            # Busca por nombre; si no, intenta por alias simples
                            fpos = self.env['account.fiscal.position'].search([('name', 'ilike', reg_raw_pos)], limit=1)
                            if not fpos:
                                pos_t = _norm(reg_raw_pos)
                                map_pos = {
                                    'exterior': 'Exterior',
                                    'export': 'Exterior',
                                    'nacional': 'Nacional',
                                    'default': 'Nacional',
                                    'exento': 'Exento',
                                }
                                alias = next((v for k, v in map_pos.items() if k in pos_t), False)
                                if alias:
                                    fpos = self.env['account.fiscal.position'].search([('name', 'ilike', alias)], limit=1)
                            if fpos and owner.property_account_position_id != fpos:
                                vals_owner['property_account_position_id'] = fpos.id

                        if vals_owner:
                            owner.sudo().write(vals_owner)
                    # ================== FIN PROPIETARIOS ==================
                        
                    # ... después de resolver prop, owner básico, etc.
                    blocks = self._iter_owner_blocks(row, headers_rev)

                    # Si esta fila es la “segunda plantilla” (solo owners/beneficiarios)
                    if self._is_owners_only_row(row, headers_rev, C):
                        self._owners_only_branch(i=i, prop=prop, blocks=blocks, lines_log=lines_log, so_name=so_name)
                        continue  # no toca contratos/lineas

                    # Si es mixta (tiene contrato Y columnas de owners), primero procesamos owners:
                    if blocks:
                        self._owners_only_branch(i=i, prop=prop, blocks=blocks, lines_log=lines_log, so_name=so_name)


                    # BENEFICIARIO (si viene)
                    benef_name = self._s(self._get(row, headers_rev, 'Beneficiario_Prop'))
                    benef_vat = self._s(self._get(row, headers_rev, 'Nro_id_Beneficiario_Prop'))
                    benef_doc_type_raw = self._s(self._get(row, headers_rev, 'Tipo_Dcto_Beneficiario_Prop'))
                    benef_ident_id = self._get_latam_ident_type_id(benef_doc_type_raw)
                    
                    if benef_name and benef_vat:
                        beneficiary = self._get_or_create_partner_by_vat(benef_vat, benef_name, ident_type_id=benef_ident_id, do_nit=False)

                # ===== VALORES SO =====
                vals = {
                    'name': so_name,
                    'ecoerp_contract': True,
                    'ecoerp_scope': scope,
                    'x_rental_start_date': f_ini or False,
                    'validity_date': f_fin or False,
                    'vigencia_meses': vig_meses or '',
                    'canon_property':canon_num,
                    'comision_inmobiliaria':True if comision_inmob > 0.0 else False,
                    'comision_inmobiliaria_porcentaje':comision_inmob or 0.0,
                    'cobro_comision_admin_ph':True if administracion_ph > 0.0 else False,
                    'administracion_ph':administracion_ph or 0.0,
                    'monto_comision_admin_ph':monto_administracion or 0.0,
                    'date_increment':date_increment or False
                }

                # 👇 partner principal según el tipo de contrato (SIN romper tu flujo)
                # - rental/alquiler -> arrendatario
                # - owner/administración (vacío o distinto de “Arriendo”) -> propietario
                # antes de crear/escribir el SO:
                partner_for_so = None
                if scope == 'rental':
                    partner_for_so = tenant
                else:
                    partner_for_so = owner
                    
                if not tenant:
                    owner1_name = _col('Nombre Propietario', 1)
                    owner1_vat  = _col('Cédula/NIT Propietario', 1)
                    owner1_doc  = _col('Tipo doc prop', 1)
                    if owner1_name and owner1_vat:
                        owner1_ident = self._get_latam_ident_type_id(owner1_doc)
                        owner_candidate = self._upsert_partner_by_vat(
                            owner1_name, owner1_vat,
                            ident_type_id=owner1_ident,
                            do_nit=bool(owner1_ident and self._is_nit_type(owner1_ident)),
                            overwrite=False
                        )
                        if owner_candidate:
                            partner_for_so = owner_candidate
                            # si quieres marcarlo como administrativo/owner:
                            if scope == 'rental':
                                scope = 'owner'
                            vals['ecoerp_scope'] = scope

                if not partner_for_so:
                    skipped += 1
                    lines_log.append(
                        f"Fila {i}: '{so_name}' ({scope}) sin partner (inquilino/propietario) → contrato OMITIDO."
                    )
                    # si quieres, igual procesa owners si vinieron:
                    if blocks:
                        self._owners_only_branch(i=i, prop=prop, blocks=blocks, lines_log=lines_log, so_name=so_name)
                    continue

                vals['partner_id'] = partner_for_so.id
                # opcional pero recomendable:
                vals.setdefault('partner_invoice_id', partner_for_so.id)
                vals.setdefault('partner_shipping_id', partner_for_so.id)


                # Mantengo lo tuyo: vincular propiedad si existe
                if prop:
                    vals['x_account_analytic_account_id'] = prop.id
                    
                # plan recurrente (solo para los que correspondan)
                plan = self.env.ref('sale_subscription.subscription_plan_month', raise_if_not_found=False)
                if plan:
                    vals['plan_id'] = plan.id
                    
                # términos de pago (elige el tuyo: por xml_id o búsqueda por nombre)
                pay_term = self.env.ref('account.account_payment_term_immediate', raise_if_not_found=False)
                if pay_term:
                    vals['payment_term_id'] = pay_term.id
                    
                existing = self._ensure_sale_order(so_name)# contrato ya existe    
                # cargamos la linea de producto de alquiler                
                # canon (renta)
                if canon_num and canon_num > 0 and not existing:
                    charge_lines = self._default_rent_charge_line(canon_num)
                    if not charge_lines:
                        raise UserError(_("No hay tarifa de alquiler en Excel o el producto no existe."))

                    vals.setdefault('order_line', [])
                    order = self  # si estás en sale.order; si estás en wizard, usa la variable del pedido

                    for ch in charge_lines:
                        amount = abs(ch.get('amount') or 0.0)
                        if float_is_zero(amount, precision_digits=2):
                            continue

                        prod = self.env['product.product'].sudo().browse(ch.get('product_id')).exists()
                        if not prod:
                            raise UserError(_("El producto de canon no existe (ID %s).") % (ch.get('product_id'),))

                        # Impuestos mapeados por posición fiscal del pedido
                        fpos = getattr(order, 'fiscal_position_id', False)
                        partner_tax = getattr(order, 'partner_shipping_id', False) or getattr(order, 'partner_id', False)
                        taxes = fpos.map_tax(prod.taxes_id, prod, partner_tax) if fpos else prod.taxes_id

                        vals['order_line'].append((0, 0, {
                            'product_id': prod.id,
                            'product_uom': prod.uom_id.id,
                            'product_uom_qty': 1.0,
                            'price_unit': amount,
                            'name': ch.get('name') or _('Canon de arrendamiento'),
                            'tax_id': [(6, 0, taxes.ids)],
                        }))

                # administración PH
                if administracion_ph > 0.0 and monto_administracion > 0.0 and not existing:
                    charge_lines = self._default_admin_charge_line(administracion_ph, monto_administracion)
                    if not charge_lines:
                        raise UserError(_("No hay administración en Excel o el producto no existe."))

                    vals.setdefault('order_line', [])
                    order = self

                    for ch in charge_lines:
                        amount = abs(ch.get('amount') or 0.0)
                        if float_is_zero(amount, precision_digits=2):
                            continue

                        prod = self.env['product.product'].sudo().browse(ch.get('product_id')).exists()
                        if not prod:
                            raise UserError(_("El producto de administración no existe (ID %s).") % (ch.get('product_id'),))

                        fpos = getattr(order, 'fiscal_position_id', False)
                        partner_tax = getattr(order, 'partner_shipping_id', False) or getattr(order, 'partner_id', False)
                        taxes = fpos.map_tax(prod.taxes_id, prod, partner_tax) if fpos else prod.taxes_id

                        vals['order_line'].append((0, 0, {
                            'product_id': prod.id,
                            'product_uom': prod.uom_id.id,
                            'product_uom_qty': 1.0,
                            'price_unit': amount,
                            'name': ch.get('name') or _('Administración PH'),
                            'tax_id': [(6, 0, taxes.ids)],
                        }))

                # ===== SIMULACIÓN =====
                if self.simulate:
                    simulated += 1
                    lines_log.append(f"Fila {i}: (SIMULADO) {so_name} [{scope}]")
                    continue

                # ===== CREATE/UPDATE =====
                ctx_skip = dict(self.env.context, skip_lessor_contact_company_check=True)
                if existing:
                    existing.sudo()\
                        .with_context(ctx_skip)\
                        .write(vals)
                    so = existing
                    updated += 1
                    lines_log.append(f"Fila {i}: actualizado {so_name}.")
                else:
                    vals.setdefault('state', 'draft')
                    so = self.env['sale.order'].sudo()\
                        .with_context(ctx_skip)\
                        .create(vals)
                    created += 1
                    lines_log.append(f"Fila {i}: creado {so_name}.")


                # >>> AQUI: guardar cláusulas del contrato
                self._upsert_clause_var(so.id, 'CONTRATO_CANON', canon_num)
                self._upsert_clause_var(so.id, 'CONTRATO_CANON_LETRAS', canon_let)

                # ===== POST: ubicación siempre sobre la propiedad =====
                if prop:
                    self._set_property_location(prop, dir_inm, mun_inm,
                            urbanizacion=self._s(self._get(row, headers_rev, 'Urbanizacion_Inmueble')),
                            barrio=self._s(self._get(row, headers_rev, 'Barrio_inmueble')),
                            tipo_prop=self._s(self._get(row, headers_rev, 'Clase')),
                            sector=self._s(self._get(row, headers_rev, 'Barrio_inmueble')))

                # Deudores 1..4 (si están las columnas). No rompe si faltan.
                guarantor_ids = []
                for k in (1, 2, 3, 4):
                    g_name = self._s(self._get(row, headers_rev, f'Nombre_coarrendatario{k}'))
                    g_vat = self._s(self._get(row, headers_rev, f'Ced_Coarrendatario{k}'))
                    g_doc_type_raw = self._s(self._get(row, headers_rev, f'Tipo_Dcto_Coarrendatario{k}'))  # si existe en tu Excel
                    g_ident_id = self._get_latam_ident_type_id(g_doc_type_raw)
                    if g_name and g_vat and 'sin coarrendatario' not in g_name.lower():
                        gp = self._get_or_create_partner_by_vat(g_vat, g_name, ident_type_id=g_ident_id, do_nit=False)
                        guarantor_ids.append(gp.id)
                    elif g_name or g_vat:
                        lines_log.append(f"Fila {i}: deudor {k} incompleto (Nombre='{g_name}', Doc='{g_vat}') → omitido.")

                if guarantor_ids and 'x_guarant_partner_id' in so._fields:
                    so.sudo().write({'x_guarant_partner_id': [(6, 0, list(set(guarantor_ids)))]})
                    
                    
                    
                # 1) Compañía base
                comp_id = prop.company_id.id if prop and prop.company_id else self.env.company.id

                # 2) Alinear compañías (si tienes estos helpers)
                self._align_company(prop, comp_id)
                self._align_company(owner, comp_id)
                self._align_company(beneficiary, comp_id)

                # 3) Limpieza opcional
                self._clear_representative(prop)
                self._clear_representative(owner)
                self._clear_representative(beneficiary)

                # # 4) Relacionar líneas con savepoints y LOG usando blocks
                for b in (blocks or []):
                    owner = b.get('owner')
                    pct   = float(b.get('pct') or 0.0)  # <- ESTE es el % correcto por propietario (incluye PROP 5)
                    if not owner:
                        continue
                    with self.env.cr.savepoint():
                        owner_line = self._link_owner_line(
                            prop=prop,
                            owner=owner,
                            comp_id=comp_id,
                            percent_pro=pct,          # <- pasar el % del bloque, NO porcent_prop
                            main_payee=False,     # propietario raíz
                        )
                    # si manejas beneficiario por bloque, llama a tu método de beneficiarios aquí:
                    beneficiary = b.get('beneficiary')
                    if beneficiary:
                        self._ensure_benef_only_create(prop=prop, owner_line=owner_line,
                                                    beneficiary=beneficiary, percent=pct)

                """ with self.env.cr.savepoint():
                    if owner_line:
                        self._link_benef_line(
                            prop=prop,
                            owner_line=owner_line,         # <- AQUÍ VA LA LÍNEA DE PROPIETARIO
                            beneficiary=beneficiary,
                            comp_id=comp_id,
                            percent=porcent_prop, # <- tu valor de beneficiario
                            main_payee=benef_name,      # <- principal del beneficiario
                        )
                    else:
                        _logger.warning("BenefLine: skip because owner_line was not created/resolved (prop=%s owner=%s)",
                                        getattr(prop, 'id', None), getattr(owner, 'id', None)) """

            if errors:
                lines_log.append("---- LOGGER ERRORES DETECTADOS ----")
                lines_log += [f"- {e}" for e in errors]
            issues = self._check_owner_percentages(lines_log) or '' or []
            lines_log += issues
            # _logger.info(" __FINAL__LOG: %s ", lines_log)
        except UserError:
            raise
        except Exception as e:
            _logger.exception("LOGGER Fallo importador Excel")
            raise UserError(_("Hubo un error al procesar el Excel: %s") % e)

        # Resumen final
        parts = [
            f"Importación finalizada. Creados={created}, Actualizados={updated}, Omitidos={skipped}, Simulados={simulated}",
            "-" * 60,
        ]
        for item in (lines_log or []):
            if item is None:
                continue
            parts.append(tools.ustr(item))

        summary_text = "\n".join(parts)

        # Guardar resumen en el wizard actual (si sigue existiendo)
        if self.exists():
            self.write({'summary': summary_text})

        # --- REGLA: si hubo omitidos/errores/alertas -> popup/modal con detalle
        def _has_issues():
            if skipped and skipped > 0:
                return True
            for x in (lines_log or []):
                s = tools.ustr(x).lower()
                if 'omitid' in s or 'omitido' in s or 'error' in s or '⚠' in s:
                    return True
            return False

        if _has_issues():
            # Intenta usar una vista de formulario del wizard si existe; si no, deja que Odoo elija
            view = self.env.ref(
                'industry_real_estate.contract_excel_import_wizard_form',  # ajusta si tu xml_id es otro
                raise_if_not_found=False
            )
            action = {
                'type': 'ir.actions.act_window',
                'name': _('Resultado de importación'),
                'res_model': self._name,
                'res_id': self.id,
                'target': 'new',          # popup/modal
                'view_mode': 'form',
                'context': dict(self.env.context or {}, show_omitted=True),
            }
            if view:
                action['views'] = [(view.id, 'form')]
                action['view_id'] = view.id
            return action

        # Si no hay issues, deja el rainbow man de éxito
        return {
            'effect': {
                'fadeout': 'slow',
                'message': _('¡Importación exitosa!'),
                'type': 'rainbow_man'
            }
        }
