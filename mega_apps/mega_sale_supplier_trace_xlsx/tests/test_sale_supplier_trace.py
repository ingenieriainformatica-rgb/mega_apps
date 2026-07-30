# -*- coding: utf-8 -*-
"""Pruebas obligatorias de mega_sale_supplier_trace_xlsx.

Cada metodo referencia, en su docstring, el punto de la lista de "PRUEBAS
OBLIGATORIAS" del encargo que cubre. Varios puntos muy relacionados se
verifican dentro de un mismo metodo cuando comparten la misma preparacion
(por ejemplo, "dos compras anteriores a la venta" y "compra anterior a
date_from" se ejercitan juntas), y se indica explicitamente cuando esto
ocurre para que el conteo final de pruebas sea verificable.

No se usa sudo() en ningun punto. No se modifica account.invoice.report.
No se usa product.supplierinfo. Todo el fixture se crea y se destruye dentro
de la transaccion de cada TransactionCase (rollback automatico).
"""
from datetime import date, timedelta

from odoo import Command, fields
from odoo.exceptions import AccessError
from odoo.tests import tagged
from odoo.tests.common import TransactionCase

from ..models.sale_supplier_trace_engine import STATUS_EXACT, STATUS_PROBABLE, STATUS_NOT_DETERMINED


@tagged('post_install', '-at_install')
class TestSaleSupplierTrace(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env.company
        cls.warehouse = cls.env['stock.warehouse'].search([('company_id', '=', cls.company.id)], limit=1)
        # Fecha centinela lejos en el futuro: esta base de pruebas es un clon
        # de un snapshot de produccion con miles de facturas reales; usar
        # "hoy" como rango del motor mezclaria esos datos reales con el
        # fixture sintetico y desbordaria el limite de rendimiento. Ninguna
        # factura real puede tener fecha en 2030.
        cls.test_date = date(2030, 6, 15)
        # country_id extranjero para que mega_contact_lock_identification no
        # exija NIT/cedula colombiana en estos contactos sinteticos de prueba.
        foreign_country = cls.env.ref('base.us')
        cls.customer = cls.env['res.partner'].create({'name': 'Cliente Trace', 'country_id': foreign_country.id})
        cls.advisor = cls.env['res.partner'].create({
            'name': 'Asesor Trace', 'country_id': foreign_country.id, 'is_advisor': True,
        })
        cls.vendor_a = cls.env['res.partner'].create({'name': 'Proveedor A Trace', 'country_id': foreign_country.id})
        cls.vendor_b = cls.env['res.partner'].create({'name': 'Proveedor B Trace', 'country_id': foreign_country.id})
        cls.vendor_c = cls.env['res.partner'].create({'name': 'Proveedor C Trace', 'country_id': foreign_country.id})

        cls.categ_fifo = cls.env['product.category'].create({
            'name': 'Trace FIFO', 'property_cost_method': 'fifo', 'property_valuation': 'real_time',
        })
        cls.uom_unit = cls.env.ref('uom.product_uom_unit')
        cls.uom_dozen = cls.env.ref('uom.product_uom_dozen')

        cls.product = cls._make_product(cls, 'Producto Trace', categ=cls.categ_fifo)
        cls.product_serial = cls._make_product(cls, 'Producto Serial Trace', categ=cls.categ_fifo,
                                                 tracking='serial')
        cls.service = cls.env['product.product'].create({
            'name': 'Servicio Trace', 'type': 'service', 'list_price': 50.0,
        })
        combo_choice = cls.env['product.combo'].create({
            'name': 'Combo Choice Trace',
            'combo_item_ids': [Command.create({'product_id': cls.service.id})],
        })
        cls.combo_product = cls.env['product.product'].create({
            'name': 'Combo Trace', 'type': 'combo', 'list_price': 10.0,
            'combo_ids': [Command.link(combo_choice.id)],
        })

        cls.engine = cls.env['sale.supplier.trace.engine']

        cls.group_user = cls.env.ref('mega_sale_supplier_trace_xlsx.group_sale_supplier_trace_user')
        cls.group_export = cls.env.ref('mega_sale_supplier_trace_xlsx.group_sale_supplier_trace_export')

    # -- helpers de fixture -----------------------------------------------

    def _make_product(self, name, categ=None, tracking='none'):
        vals = {
            'name': name, 'type': 'consu', 'is_storable': True, 'list_price': 100.0,
            'tracking': tracking,
        }
        if categ:
            vals['categ_id'] = categ.id
        return self.env['product.product'].create(vals)

    def _buy(self, product, vendor, qty, price=100.0, lot_name=None):
        """Crea, confirma y recibe una orden de compra. Devuelve (po, receipt_move)."""
        po = self.env['purchase.order'].create({
            'partner_id': vendor.id,
            'order_line': [Command.create({
                'product_id': product.id, 'product_qty': qty, 'product_uom': product.uom_id.id,
                'price_unit': price, 'date_planned': fields.Datetime.now(),
            })],
        })
        po.button_confirm()
        picking = po.picking_ids.filtered(lambda p: p.state != 'done')[:1]
        picking.action_assign()
        move = picking.move_ids[:1]
        move.quantity = qty
        move.picked = True
        if lot_name and product.tracking != 'none':
            self.env['stock.lot'].create({'name': lot_name, 'product_id': product.id, 'company_id': self.company.id})
            move.move_line_ids.lot_id = self.env['stock.lot'].search(
                [('name', '=', lot_name), ('product_id', '=', product.id)], limit=1)
        picking._action_done()
        return po, move

    def _sell(self, product, customer, qty, price=150.0, invoice_policy='order'):
        product.invoice_policy = invoice_policy
        so = self.env['sale.order'].create({
            'partner_id': customer.id,
            'warehouse_id': self.warehouse.id,
            'advisor_id': self.advisor.id,
            'order_line': [Command.create({'product_id': product.id, 'product_uom_qty': qty, 'price_unit': price})],
        })
        so.action_confirm()
        return so

    def _deliver(self, so, qty=None, cancel_backorder=True):
        picking = so.picking_ids.filtered(lambda p: p.state not in ('done', 'cancel'))[:1]
        picking.action_assign()
        move = picking.move_ids[:1]
        move.quantity = qty if qty is not None else move.product_uom_qty
        move.picked = True
        ctx = {'cancel_backorder': True} if cancel_backorder else {}
        picking.with_context(**ctx)._action_done()
        return move

    def _invoice(self, so, post=True):
        invoice = so._create_invoices()
        invoice.invoice_date = self.test_date
        if post:
            invoice.action_post()
        return invoice

    def _credit_note(self, invoice, qty=None):
        move_reversal = self.env['account.move.reversal'].with_context(
            active_model='account.move', active_ids=invoice.ids).create({
                'journal_id': invoice.journal_id.id,
                'reason': 'Prueba trazabilidad',
            })
        result = move_reversal.reverse_moves()
        credit_note = self.env['account.move'].browse(result['res_id'])
        if qty is not None:
            credit_note.invoice_line_ids.filtered(lambda l: l.display_type == 'product').quantity = qty
        credit_note.invoice_date = self.test_date
        credit_note.action_post()
        return credit_note

    def _run(self, **overrides):
        params = {
            'date_from': self.test_date.replace(day=1),
            'date_to': self.test_date + timedelta(days=15),
            'partner_id': None, 'product_id': None, 'product_categ_id': None,
            'move_id': None, 'supplier_id': None, 'journal_id': None,
            'doc_type': 'both', 'include_services': True,
            'company_ids': [self.company.id],
        }
        params.update(overrides)
        return self.engine.run(params)

    def _rows_for_move(self, result, move):
        return [r for r in result['all_rows'] if r['move'] == move]

    # -- 1. Trazabilidad exacta por lote/serie -----------------------------

    def test_01_lot_serial_exact(self):
        """Punto 4 / lista nueva: lote-serie con recepcion+PO -> exacto."""
        po, receipt = self._buy(self.product_serial, self.vendor_a, 1, lot_name='LOT-001')
        so = self._sell(self.product_serial, self.customer, 1)
        move = self._deliver(so)
        move.move_line_ids.lot_id = receipt.move_line_ids.lot_id
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['status'], STATUS_EXACT)
        self.assertEqual(rows[0]['supplier'], self.vendor_a)
        self.assertAlmostEqual(rows[0]['qty'], 1.0)

    # -- 2. FIFO probable, nunca exacto, aunque el producto sea FIFO -------

    def test_02_fifo_never_exact(self):
        """Puntos 2 y 3: reconstruccion cronologica -> siempre 'probable'."""
        self._buy(self.product, self.vendor_a, 5, price=100.0)
        so = self._sell(self.product, self.customer, 5)
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertTrue(rows)
        for r in rows:
            self.assertEqual(r['status'], STATUS_PROBABLE)
            self.assertNotEqual(r['status'], STATUS_EXACT)
        self.assertEqual(rows[0]['supplier'], self.vendor_a)

    # -- 3/4. Compra anterior a date_from + dos compras anteriores a la venta --

    def test_03_purchases_before_date_from_two_suppliers(self):
        """Lista nueva #2 y #3: el balance antes de date_from se reconstruye
        con TODO el historial, no solo desde date_from."""
        self._buy(self.product, self.vendor_a, 3, price=90.0)
        self._buy(self.product, self.vendor_b, 3, price=110.0)
        so = self._sell(self.product, self.customer, 4)
        self._deliver(so)
        invoice = self._invoice(so)

        # date_from posterior a ambas compras: el motor igual debe reconstruir
        # el consumo correctamente porque replica TODO el historial del producto.
        result = self._run(date_from=self.test_date)
        rows = self._rows_for_move(result, invoice)
        total_qty = sum(r['qty'] for r in rows)
        self.assertAlmostEqual(total_qty, 4.0, places=4)
        suppliers = {r['supplier'] for r in rows}
        self.assertEqual(suppliers, {self.vendor_a, self.vendor_b})
        qty_a = sum(r['qty'] for r in rows if r['supplier'] == self.vendor_a)
        qty_b = sum(r['qty'] for r in rows if r['supplier'] == self.vendor_b)
        self.assertAlmostEqual(qty_a, 3.0, places=4)  # FIFO: se agota primero el proveedor A
        self.assertAlmostEqual(qty_b, 1.0, places=4)

    # -- 5. Inventario inicial sin proveedor (ajuste positivo antes de vender) --

    def test_04_opening_inventory_without_purchase(self):
        """Lista nueva #1: una capa de apertura/ajuste sin purchase_line_id
        consumida por una venta debe quedar No determinado, nunca asignar
        un proveedor por descarte."""
        quant = self.env['stock.quant'].with_context(inventory_mode=True).create({
            'product_id': self.product.id, 'location_id': self.warehouse.lot_stock_id.id,
            'inventory_quantity': 2,
        })
        quant.action_apply_inventory()
        so = self._sell(self.product, self.customer, 2)
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertTrue(all(r['status'] == STATUS_NOT_DETERMINED for r in rows))
        self.assertTrue(all(not r['supplier'] for r in rows))

    # -- 6. Negative stock + regularizacion posterior -----------------------

    def test_05_negative_stock_then_regularized(self):
        """Lista nueva #8 y #9: venta sin stock previo -> No determinado; una
        compra POSTERIOR no debe atribuirse retroactivamente a esa venta."""
        so = self._sell(self.product, self.customer, 2)
        self._deliver(so)  # sin compra previa: inventario negativo
        invoice = self._invoice(so)

        po2, _receipt2 = self._buy(self.product, self.vendor_c, 5, price=120.0)
        so2 = self._sell(self.product, self.customer, 1)
        self._deliver(so2)
        invoice2 = self._invoice(so2)

        result = self._run()
        rows1 = self._rows_for_move(result, invoice)
        self.assertTrue(all(r['status'] == STATUS_NOT_DETERMINED for r in rows1))

        rows2 = self._rows_for_move(result, invoice2)
        self.assertTrue(any(r['status'] == STATUS_PROBABLE and r['supplier'] == self.vendor_c for r in rows2))

    # -- 7. Transferencia interna: no debe generar capa de valoracion -------

    def test_06_internal_transfer_not_a_candidate(self):
        """Lista nueva #11: una transferencia interna (internal->internal) no
        crea stock.valuation.layer y por lo tanto no puede aparecer como
        proveedor de nada."""
        self._buy(self.product, self.vendor_a, 5, price=100.0)
        other_location = self.env['stock.location'].create({
            'name': 'Trace Secondary', 'usage': 'internal',
            'location_id': self.warehouse.view_location_id.id,
        })
        internal_type = self.env['stock.picking.type'].search(
            [('code', '=', 'internal'), ('warehouse_id', '=', self.warehouse.id)], limit=1)
        transfer = self.env['stock.picking'].create({
            'picking_type_id': internal_type.id,
            'location_id': self.warehouse.lot_stock_id.id,
            'location_dest_id': other_location.id,
            'move_ids': [Command.create({
                'name': self.product.name, 'product_id': self.product.id, 'product_uom_qty': 2,
                'product_uom': self.product.uom_id.id,
                'location_id': self.warehouse.lot_stock_id.id, 'location_dest_id': other_location.id,
            })],
        })
        transfer.action_confirm()
        transfer.action_assign()
        transfer.move_ids.quantity = 2
        transfer.move_ids.picked = True
        transfer._action_done()

        svl_count = self.env['stock.valuation.layer'].search_count([
            ('stock_move_id', 'in', transfer.move_ids.ids),
        ])
        self.assertEqual(svl_count, 0, 'Una transferencia interna no debe generar stock.valuation.layer')

        so = self._sell(self.product, self.customer, 5)
        self._deliver(so)
        invoice = self._invoice(so)
        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertTrue(all(r['supplier'] in (False, self.vendor_a) for r in rows))
        self.assertAlmostEqual(sum(r['qty'] for r in rows if r['supplier'] == self.vendor_a), 5.0, places=4)

    # -- 8. Scrap: consume la cola FIFO pero no se atribuye a ninguna venta --

    def test_07_scrap_consumes_pool_not_a_sale(self):
        """Lista nueva #12: el scrap consume capas FIFO en la simulacion (para
        no descuadrar el saldo) pero nunca genera una fila de venta."""
        self._buy(self.product, self.vendor_a, 5, price=100.0)
        self.env['stock.scrap'].create({
            'product_id': self.product.id, 'scrap_qty': 2,
            'location_id': self.warehouse.lot_stock_id.id,
        }).do_scrap()

        so = self._sell(self.product, self.customer, 3)
        self._deliver(so)
        invoice = self._invoice(so)
        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertAlmostEqual(sum(r['qty'] for r in rows), 3.0, places=4)
        self.assertTrue(all(r['supplier'] == self.vendor_a for r in rows))

    # -- 9/10. Devolucion de cliente y nota credito sin/con devolucion fisica --

    def test_08_customer_return_and_credit_note(self):
        """Lista nueva #13, #15, #16: nota credito hereda proporcionalmente
        de la factura original, nunca consulta el proveedor actual del
        producto ni ejecuta una nueva reconstruccion FIFO."""
        self._buy(self.product, self.vendor_a, 5, price=100.0)
        so = self._sell(self.product, self.customer, 4)
        out_move = self._deliver(so)
        invoice = self._invoice(so)

        # nota credito SIN devolucion fisica (reversion contable pura)
        credit_no_return = self._credit_note(invoice, qty=1)

        # devolucion fisica real de 1 unidad + su nota credito
        return_move = out_move.copy({
            'location_id': out_move.location_dest_id.id,
            'location_dest_id': out_move.location_id.id,
            'origin_returned_move_id': out_move.id,
            'product_uom_qty': 1,
        })
        return_move._action_confirm()
        return_move._action_assign()
        return_move.quantity = 1
        return_move.picked = True
        return_move._action_done()

        result = self._run()
        original_rows = self._rows_for_move(result, invoice)
        credit_rows = self._rows_for_move(result, credit_no_return)

        self.assertTrue(all(r['qty'] < 0 for r in credit_rows), 'La nota credito debe representarse en negativo')
        self.assertAlmostEqual(sum(-r['qty'] for r in credit_rows), 1.0, places=4)
        self.assertEqual({r['supplier'] for r in credit_rows}, {r['supplier'] for r in original_rows})
        for r in credit_rows:
            self.assertIn('Heredado proporcionalmente', r['note'])

    # -- 11. Devolucion al proveedor: consume la cola sin ser una venta -----

    def test_09_return_to_vendor(self):
        """Lista nueva #14: una devolucion al proveedor consume la cola FIFO
        (para que la contabilidad cuadre) pero no debe aparecer como una
        fila de venta ni alterar el proveedor de ventas posteriores."""
        po, receipt = self._buy(self.product, self.vendor_a, 5, price=100.0)
        return_to_vendor = receipt.copy({
            'location_id': receipt.location_dest_id.id,
            'location_dest_id': receipt.location_id.id,
            'origin_returned_move_id': receipt.id,
            'product_uom_qty': 1,
        })
        return_to_vendor._action_confirm()
        return_to_vendor._action_assign()
        return_to_vendor.quantity = 1
        return_to_vendor.picked = True
        return_to_vendor._action_done()

        so = self._sell(self.product, self.customer, 3)
        self._deliver(so)
        invoice = self._invoice(so)
        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertAlmostEqual(sum(r['qty'] for r in rows), 3.0, places=4)
        self.assertTrue(all(r['supplier'] == self.vendor_a for r in rows))

    # -- 12. Facturacion parcial: dos facturas sobre una misma entrega ------

    def test_10_two_invoices_on_one_delivery(self):
        """Lista nueva #5 y #6: la misma entrega no debe atribuirse completa
        a cada factura (evita duplicar cantidades)."""
        self._buy(self.product, self.vendor_a, 10, price=100.0)
        product = self.product
        product.invoice_policy = 'delivery'
        so = self.env['sale.order'].create({
            'partner_id': self.customer.id,
            'warehouse_id': self.warehouse.id,
            'advisor_id': self.advisor.id,
            'order_line': [Command.create({'product_id': product.id, 'product_uom_qty': 6, 'price_unit': 150.0})],
        })
        so.action_confirm()
        self._deliver(so)

        so.order_line.qty_delivered = 6
        invoice1 = so.with_context(default_invoice_policy='delivery')._create_invoices()
        invoice1.invoice_line_ids.filtered(lambda l: l.display_type == 'product').quantity = 4
        invoice1.invoice_date = self.test_date
        invoice1.action_post()

        invoice2 = so._create_invoices()
        invoice2.invoice_line_ids.filtered(lambda l: l.display_type == 'product').quantity = 2
        invoice2.invoice_date = self.test_date
        invoice2.action_post()

        result = self._run()
        rows1 = self._rows_for_move(result, invoice1)
        rows2 = self._rows_for_move(result, invoice2)
        self.assertAlmostEqual(sum(r['qty'] for r in rows1), 4.0, places=4)
        self.assertAlmostEqual(sum(r['qty'] for r in rows2), 2.0, places=4)
        # ninguna de las dos reclama las 6 unidades completas de la entrega
        self.assertNotAlmostEqual(sum(r['qty'] for r in rows1), 6.0, places=4)

    # -- 13. Dos entregas (backorder) sobre una sola factura -----------------

    def test_11_two_deliveries_one_invoice(self):
        """Lista nueva #7."""
        self._buy(self.product, self.vendor_a, 2, price=90.0)
        self._buy(self.product, self.vendor_b, 2, price=110.0)
        so = self._sell(self.product, self.customer, 4, invoice_policy='order')
        self._deliver(so, qty=2, cancel_backorder=False)  # backorder por las 2 restantes
        backorder = so.picking_ids.filtered(lambda p: p.state not in ('done', 'cancel'))
        backorder.action_assign()
        backorder.move_ids.quantity = 2
        backorder.move_ids.picked = True
        backorder.with_context(cancel_backorder=True)._action_done()
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertAlmostEqual(sum(r['qty'] for r in rows), 4.0, places=4)
        delivery_moves = {r['delivery_move'] for r in rows}
        self.assertEqual(len(delivery_moves), 2, 'Debe repartirse entre las dos entregas, no colapsarlas')

    # -- 14. Venta repartida entre proveedores + redondeo con tres proveedores --

    def test_12_three_suppliers_rounding(self):
        """Lista nueva #4 y #18: la suma de subtotales atribuidos debe
        coincidir EXACTAMENTE con el subtotal original, con el residuo de
        redondeo aplicado solo dentro de la misma linea."""
        self._buy(self.product, self.vendor_a, 1, price=33.33)
        self._buy(self.product, self.vendor_b, 1, price=33.33)
        self._buy(self.product, self.vendor_c, 1, price=33.34)
        so = self._sell(self.product, self.customer, 3, price=100.0)
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertEqual({r['supplier'] for r in rows}, {self.vendor_a, self.vendor_b, self.vendor_c})
        line = invoice.invoice_line_ids.filtered(lambda l: l.display_type == 'product')
        self.assertEqual(round(sum(r['subtotal'] for r in rows), 2), round(line.price_subtotal, 2))
        self.assertEqual(round(sum(r['tax'] for r in rows), 2), round(line.price_total - line.price_subtotal, 2))
        self.assertEqual(round(sum(r['total'] for r in rows), 2), round(line.price_total, 2))

    # -- 15. Distintas unidades de medida ------------------------------------

    def test_13_different_uom(self):
        """Lista nueva #17: la conversion debe usar uom.uom._compute_quantity,
        y las cantidades deben cuadrar en la unidad reportada."""
        self._buy(self.product, self.vendor_a, 24, price=10.0)  # 24 unidades
        so = self.env['sale.order'].create({
            'partner_id': self.customer.id,
            'warehouse_id': self.warehouse.id,
            'advisor_id': self.advisor.id,
            'order_line': [Command.create({
                'product_id': self.product.id, 'product_uom_qty': 1,
                'product_uom': self.uom_dozen.id, 'price_unit': 120.0,
            })],
        })
        so.action_confirm()
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        total_in_report_uom = sum(r['qty'] for r in rows)
        # las filas se reportan en la UoM de la linea de factura (docena)
        self.assertEqual(rows[0]['uom'], self.uom_dozen)
        self.assertAlmostEqual(total_in_report_uom, 1.0, places=4)

    # -- 16. Linea con cantidad cero / seccion o nota / combo ---------------

    def test_14_zero_qty_section_combo_lines_are_harmless(self):
        """Lista nueva #19, #20, #21. Se factura directamente por account.move
        (sin pasar por el flujo de seleccion de combo de ventas, que expande
        la linea de forma independiente al motor de trazabilidad) para
        aislar unicamente el comportamiento del motor ante estas lineas."""
        invoice = self.env['account.move'].create({
            'move_type': 'out_invoice', 'partner_id': self.customer.id, 'invoice_date': self.test_date,
            'invoice_line_ids': [
                Command.create({'display_type': 'line_section', 'name': 'Seccion de prueba'}),
                Command.create({'product_id': self.combo_product.id, 'quantity': 1, 'price_unit': 10.0}),
                Command.create({'product_id': self.product.id, 'quantity': 0.0001, 'price_unit': 100.0}),
            ],
        })
        invoice.action_post()
        # no debe reventar: ni la seccion (display_type != 'product', excluida
        # del dominio), ni el combo (no almacenable), ni una cantidad casi
        # cero de un producto sin compra previa deben producir excepciones.
        result = self._run()
        rows = self._rows_for_move(result, invoice)
        self.assertTrue(rows)
        self.assertTrue(all(r['status'] == STATUS_NOT_DETERMINED for r in rows))

    # -- 17/18/19. Seguridad: consulta, exportacion, sin grupos, menu --------

    def test_15_security_groups(self):
        """Lista nueva #22, #23, #24, #25."""
        # El asistente lee stock.valuation.layer y purchase.order.line: ademas
        # del grupo propio, el usuario necesita acceso de lectura basico a
        # Inventario y Compras (esto NO lo otorga nuestro grupo -- el motor
        # nunca usa sudo() -- es un prerequisito de acceso que se documenta).
        base_groups = [
            self.env.ref('account.group_account_invoice').id,
            self.env.ref('stock.group_stock_user').id,
            self.env.ref('purchase.group_purchase_user').id,
        ]
        user_consult = self.env['res.users'].create({
            'name': 'Usuario Consulta', 'login': 'trace_user_consult',
            'groups_id': [Command.link(self.group_user.id)] + [Command.link(g) for g in base_groups],
        })
        user_export = self.env['res.users'].create({
            'name': 'Usuario Exporta', 'login': 'trace_user_export',
            'groups_id': [Command.link(self.group_export.id)] + [Command.link(g) for g in base_groups],
        })
        user_none = self.env['res.users'].create({
            'name': 'Usuario Sin Grupos', 'login': 'trace_user_none',
            'groups_id': [Command.link(self.env.ref('base.group_user').id)],
        })

        wizard_consult = self.env['sale.supplier.trace.wizard'].with_user(user_consult).create({
            'date_from': self.test_date.replace(day=1), 'date_to': self.test_date + timedelta(days=15),
        })
        wizard_consult.action_preview()  # no debe lanzar AccessError
        with self.assertRaises(AccessError):
            wizard_consult.action_generate_xlsx()

        wizard_export = self.env['sale.supplier.trace.wizard'].with_user(user_export).create({
            'date_from': self.test_date.replace(day=1), 'date_to': self.test_date + timedelta(days=15),
        })
        wizard_export.action_generate_xlsx()  # no debe lanzar AccessError
        self.assertTrue(wizard_export.xlsx_file)

        with self.assertRaises(AccessError):
            self.env['sale.supplier.trace.wizard'].with_user(user_none).create({
                'date_from': self.test_date, 'date_to': self.test_date,
            })

        menu = self.env.ref('mega_sale_supplier_trace_xlsx.menu_sale_supplier_trace_wizard')
        visible_consult = self.env['ir.ui.menu'].with_user(user_consult).load_menus(False)
        visible_none = self.env['ir.ui.menu'].with_user(user_none).load_menus(False)
        self.assertIn(menu.id, visible_consult, 'El usuario con el grupo de consulta debe ver el menu')
        self.assertNotIn(menu.id, visible_none, 'Un usuario sin grupos no debe ver el menu')

        # los grupos propios no deben otorgar permisos contables adicionales
        with self.assertRaises(AccessError):
            self.env['account.move'].with_user(user_none).search([], limit=1)

    # -- 20. Archivo XLSX real: openpyxl, celdas numericas y fechas ---------

    def test_16_xlsx_file_structure(self):
        """Lista nueva #26, #27, #28."""
        import io
        import base64
        from openpyxl import load_workbook

        self._buy(self.product, self.vendor_a, 3, price=100.0)
        so = self._sell(self.product, self.customer, 3)
        self._deliver(so)
        self._invoice(so)

        self.env.user.write({'groups_id': [Command.link(self.group_export.id)]})
        wizard = self.env['sale.supplier.trace.wizard'].create({
            'date_from': self.test_date.replace(day=1), 'date_to': self.test_date + timedelta(days=15),
        })
        wizard.action_generate_xlsx()
        self.assertTrue(wizard.xlsx_file)
        content = base64.b64decode(wizard.xlsx_file)
        wb = load_workbook(io.BytesIO(content), data_only=False)
        self.assertEqual(
            set(wb.sheetnames),
            {'Detalle', 'Resumen por proveedor', 'No determinados', 'Parametros y advertencias'})

        detail = wb['Detalle']
        header = [c.value for c in next(detail.iter_rows(min_row=1, max_row=1))]
        self.assertIn('Cantidad atribuida', header)
        self.assertIn('Subtotal atribuido', header)

        # primera fila de datos: cantidad numerica, fecha real (no texto), monto numerico
        data_row = next(detail.iter_rows(min_row=2, max_row=2))
        qty_cell = data_row[header.index('Cantidad atribuida')]
        date_cell = data_row[header.index('Fecha de factura')]
        amount_cell = data_row[header.index('Total atribuido')]
        self.assertIsInstance(qty_cell.value, (int, float))
        self.assertIsInstance(amount_cell.value, (int, float))
        import datetime
        self.assertIsInstance(date_cell.value, (datetime.date, datetime.datetime))

    # -- 21. Totales sin diferencias (sin filtro de proveedor) ---------------

    def test_17_totals_match_without_supplier_filter(self):
        """Lista nueva #29."""
        self._buy(self.product, self.vendor_a, 4, price=100.0)
        so = self._sell(self.product, self.customer, 4, price=150.0)
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        total_original = sum(result['move_lines'].mapped('price_total'))
        total_attributed = sum(r['total'] for r in result['rows'])
        self.assertAlmostEqual(total_original, total_attributed, places=2)

    # -- 22/23. account.invoice.report y modelos contables intactos ---------

    def test_18_no_writes_on_core_models(self):
        """Lista nueva #30 y #31: se compara un snapshot antes/despues de
        correr el motor sobre facturas ya existentes; nada debe cambiar."""
        self._buy(self.product, self.vendor_a, 3, price=100.0)
        so = self._sell(self.product, self.customer, 3)
        self._deliver(so)
        invoice = self._invoice(so)

        report_before = self.env['account.invoice.report'].search_read(
            [('move_id', '=', invoice.id)], ['quantity', 'price_subtotal', 'price_total'])
        move_write_date_before = invoice.write_date
        line_write_dates_before = {l.id: l.write_date for l in invoice.line_ids}
        svl_count_before = self.env['stock.valuation.layer'].search_count([])

        self._run()  # ejecuta el motor completo, incluida la reconstruccion FIFO

        report_after = self.env['account.invoice.report'].search_read(
            [('move_id', '=', invoice.id)], ['quantity', 'price_subtotal', 'price_total'])
        self.assertEqual(report_before, report_after)
        self.assertEqual(invoice.write_date, move_write_date_before)
        for line in invoice.line_ids:
            self.assertEqual(line.write_date, line_write_dates_before[line.id])
        self.assertEqual(self.env['stock.valuation.layer'].search_count([]), svl_count_before)

    # -- 24. Filtros del asistente (fecha, cliente, producto, proveedor) ----

    def test_19_wizard_filters(self):
        """Lista original: filtros por fecha, cliente, producto, proveedor."""
        self._buy(self.product, self.vendor_a, 3, price=100.0)
        other_customer = self.env['res.partner'].create(
            {'name': 'Otro Cliente Trace', 'country_id': self.env.ref('base.us').id})
        so1 = self._sell(self.product, self.customer, 2)
        self._deliver(so1)
        invoice1 = self._invoice(so1)
        so2 = self._sell(self.product, other_customer, 1)
        self._deliver(so2)
        invoice2 = self._invoice(so2)

        result_partner = self._run(partner_id=self.customer.id)
        self.assertTrue(all(l.move_id.partner_id == self.customer for l in result_partner['move_lines']))

        result_product = self._run(product_id=self.product.id)
        self.assertTrue(all(l.product_id == self.product for l in result_product['move_lines']))

        result_supplier = self._run(supplier_id=self.vendor_a.id)
        self.assertTrue(all(r['supplier'] == self.vendor_a for r in result_supplier['rows']))

        old_result = self._run(date_from=date(2000, 1, 1), date_to=date(2000, 12, 31))
        self.assertEqual(len(old_result['move_lines']), 0)

    # -- 25. Rendimiento: limite configurable con error claro -----------------

    def test_20_performance_limit_raises_clear_error(self):
        """Punto 17: el limite se lee de ir.config_parameter y produce un
        error explicito, no un Excel incompleto en silencio."""
        self.env['ir.config_parameter'].set_param('mega_sale_supplier_trace_xlsx.max_lines', '0')
        self._buy(self.product, self.vendor_a, 1, price=100.0)
        so = self._sell(self.product, self.customer, 1)
        self._deliver(so)
        self._invoice(so)
        from odoo.exceptions import UserError
        with self.assertRaises(UserError):
            self._run()
        self.env['ir.config_parameter'].set_param('mega_sale_supplier_trace_xlsx.max_lines', '5000')

    # -- 26. product.supplierinfo jamas se usa como fuente de evidencia -----

    def test_21_supplierinfo_never_used(self):
        """Punto 18: aunque el producto tenga product.supplierinfo con OTRO
        proveedor configurado, el motor debe ignorarlo por completo."""
        self.env['product.supplierinfo'].create({
            'partner_id': self.vendor_b.id, 'product_tmpl_id': self.product.product_tmpl_id.id,
            'sequence': 1, 'min_qty': 0, 'price': 1.0,
        })
        self._buy(self.product, self.vendor_a, 3, price=100.0)  # compra real: proveedor A
        so = self._sell(self.product, self.customer, 3)
        self._deliver(so)
        invoice = self._invoice(so)

        result = self._run()
        rows = self._rows_for_move(result, invoice)
        # A pesar de que product.supplierinfo apunta al proveedor B con sequence=1,
        # el resultado debe ser el proveedor A (evidencia operativa real).
        self.assertTrue(all(r['supplier'] == self.vendor_a for r in rows))
        self.assertFalse(any(r['supplier'] == self.vendor_b for r in rows))
