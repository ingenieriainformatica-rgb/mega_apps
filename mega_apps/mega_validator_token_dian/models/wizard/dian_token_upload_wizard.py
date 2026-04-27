# -*- coding: utf-8 -*-
import base64
import logging
import os
from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook  # type: ignore

from odoo import fields, models, _  # type: ignore
from odoo.exceptions import UserError  # type: ignore

_logger = logging.getLogger(__name__)

ALLOWED_DOCUMENT_TYPES = {
    "Factura electrónica",
    "Nota de crédito electrónica",
    "Documento soporte con no obligados",
}

EXCLUDED_NIT_EMISOR = "900325964"


class MegaDianTokenUploadWizard(models.TransientModel):
    _name = "mega.dian.token.upload.wizard"
    _description = "Wizard subir archivo DIAN"

    token_id = fields.Many2one(
        "mega.dian.token",
        string="Token DIAN",
        required=True,
    )
    file_name = fields.Char(string="Nombre archivo")
    file_data = fields.Binary(string="Archivo", attachment=True)

    def action_confirm(self):
        self.ensure_one()

        self._validate_upload_request()

        worksheet = self._get_active_worksheet()
        headers = self._get_headers(worksheet)

        created_lines = self._create_valid_token_lines(worksheet, headers)

        if not created_lines:
            raise UserError(_(
                "El archivo no generó registros válidos. "
                "El token permanecerá en su estado actual."
            ))

        self.token_id.state = "loaded"

        return {
            "type": "ir.actions.client",
            "tag": "reload",
        }

    # ============================================================
    # VALIDACIONES
    # ============================================================

    def _validate_upload_request(self):
        if not self.token_id:
            raise UserError(_("No se encontró el token DIAN relacionado."))

        if not self.file_data:
            raise UserError(_("Debes cargar un archivo."))

        if not self.file_name:
            raise UserError(_("Debes indicar el nombre del archivo."))

        extension = os.path.splitext(self.file_name.lower())[1]
        if extension != ".xlsx":
            raise UserError(_("Solo se permite cargar archivos Excel con extensión .xlsx."))

    # ============================================================
    # LECTURA DE EXCEL
    # ============================================================

    def _get_active_worksheet(self):
        try:
            excel_bytes = base64.b64decode(self.file_data)
            workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
            return workbook.active
        except Exception as error:
            _logger.exception("Error leyendo archivo Excel del token DIAN.")
            raise UserError(_(
                "No fue posible leer el archivo Excel. "
                "Verifica que el archivo no esté dañado y tenga formato .xlsx."
            )) from error

    def _get_headers(self, worksheet):
        headers = []
        for cell in worksheet[1]:  # type: ignore
            headers.append(self._normalize_cell_value(cell.value))

        return headers

    # ============================================================
    # PROCESAMIENTO DE FILAS
    # ============================================================

    def _create_valid_token_lines(self, worksheet, headers):
        created_count = 0
        token_file_model = self.env["mega.dian.token.file"]

        for row in worksheet.iter_rows(min_row=2, values_only=True):  # type: ignore
            row_data = dict(zip(headers, row))
            parsed_row = self._parse_row_data(row_data)

            if not self._is_valid_row(parsed_row):
                continue

            token_file_model.create({
                "token_id": self.token_id.id,
                "tipo_documento": parsed_row["tipo_documento"],
                "folio": parsed_row["folio"],
                "prefijo": parsed_row["prefijo"],
                "nit_emisor": parsed_row["nit_emisor"],
                "fecha_emision_dian": parsed_row["fecha_emision_dian"],
                "nombre_emisor": parsed_row["nombre_emisor"],
                "iva": parsed_row["iva"],
                "total": parsed_row["total"],
                "cufe": self._normalize_text(row_data.get("CUFE/CUDE")),
            })
            created_count += 1

        return created_count


    def _parse_excel_date(self, value):
        if not value:
            return False

        if hasattr(value, "date"):
            return value.date()

        if isinstance(value, str):
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue

        return False

    def _parse_row_data(self, row_data):
        tipo_documento = self._normalize_text(row_data.get("Tipo de documento"))
        nit_emisor = self._normalize_text(row_data.get("NIT Emisor"))

        return {
            "tipo_documento": tipo_documento,
            "folio": self._normalize_text(row_data.get("Folio")),
            "prefijo": self._normalize_text(row_data.get("Prefijo")),
            "nit_emisor": nit_emisor,
            "nombre_emisor": self._normalize_text(row_data.get("Nombre Emisor")),
            "fecha_emision_dian": self._parse_excel_date(row_data.get("Fecha Emisión")),
            "iva": self._to_float(row_data.get("IVA")),
            "total": self._to_float(row_data.get("Total")),
        }

    def _is_valid_row(self, parsed_row):
        return (
            parsed_row["tipo_documento"] in ALLOWED_DOCUMENT_TYPES
            and parsed_row["nit_emisor"] != EXCLUDED_NIT_EMISOR
        )

    # ============================================================
    # HELPERS
    # ============================================================

    def _normalize_cell_value(self, value):
        if isinstance(value, str):
            return value.strip()
        return value or ""

    def _normalize_text(self, value):
        return str(value).strip() if value not in (None, False, "") else ""

    def _to_float(self, value):
        return float(value or 0.0)
