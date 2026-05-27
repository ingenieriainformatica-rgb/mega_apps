# -*- coding: utf-8 -*-

import base64
import csv
import io
import logging
import re
import unicodedata

from odoo import fields, models, _  #type:ignore
from odoo.exceptions import UserError  #type:ignore

_logger = logging.getLogger(__name__)


class MegaBatteryPriceImportWizard(models.TransientModel):
    _name = "mega.battery.price.import.wizard"
    _description = "Importar precios de baterías MAC"

    file = fields.Binary(string="Archivo", required=True)
    filename = fields.Char(string="Nombre del archivo")

    update_description = fields.Boolean(string="Actualizar descripción", default=True)
    update_stock = fields.Boolean(string="Actualizar existencias", default=True)
    update_prices = fields.Boolean(string="Actualizar precios", default=True)

    clear_existing_prices = fields.Boolean(
        string="Limpiar precios actuales antes de importar",
        default=True,
        help="Si está activo, se pondrán en cero los precios actuales antes de cargar el archivo.",
    )

    summary = fields.Text(string="Resumen", readonly=True)

    def action_import_prices(self):
        self.ensure_one()

        rows = self._read_file()
        if not rows:
            raise UserError(_("El archivo no contiene filas para importar."))

        Option = self.env["mega.battery.application.option"].sudo()
        currency = self.env.company.currency_id

        if self.clear_existing_prices:
            Option.search([]).write({
                "sale_price": 0.0,
                "min_sale_price": 0.0,
                "max_sale_price": 0.0,
            })

        total_rows = updated_references = updated_options = not_found = skipped = 0
        not_found_refs = []

        for row_number, row in rows:
            total_rows += 1

            description = self._get_value(row, ["descripcion", "description"])
            new_reference = self._normalize_reference(
                self._get_value(row, ["referencia", "reference"])
            )
            old_reference = self._normalize_reference(
                self._get_value(row, ["referencia_anterior", "referencia_vieja", "old_reference"])
            )
            brand_line = self._normalize_brand_line(
                self._get_value(row, ["marca", "linea", "línea"])
            )

            if not new_reference and description:
                new_reference = self._normalize_reference(self._extract_reference(description))

            if not new_reference and not old_reference:
                skipped += 1
                continue

            options = self._find_options_for_price_import(
                Option,
                new_reference=new_reference,
                old_reference=old_reference,
                brand_line=brand_line,
            )

            if not options:
                not_found += 1
                not_found_refs.append(
                    f"Fila {row_number}: nueva={new_reference or '-'} "
                    f"anterior={old_reference or '-'} marca={brand_line or '-'}"
                )
                continue

            vals = {"currency_id": currency.id}

            if new_reference:
                vals["reference"] = new_reference

            if old_reference and "old_reference" in Option._fields:
                vals["old_reference"] = old_reference

            if self.update_description:
                vals.update({
                    "description": description,
                    "uom_name": self._get_value(row, ["um", "unidad", "u_m"]),
                })

            if self.update_stock:
                vals["stock_qty"] = self._parse_number(
                    self._get_value(row, ["existencias", "stock"])
                )

            if self.update_prices:
                average_cost = self._parse_money(self._get_value(row, ["promedio"]))
                tax_amount = self._parse_money(self._get_value(row, ["iva"]))
                cost_with_tax = self._parse_money(
                    self._get_value(row, [
                        "costo_iva",
                        "costo_mas_iva",
                        "costo_con_iva",
                        "costo_iva_moneda",
                        "costo_total",
                    ])
                )

                min_sale_price = self._get_min_sale_price_from_file(row)
                sale_price = self._get_full_sale_price_from_file(row)

                if not min_sale_price:
                    min_sale_price = self._calculate_sale_price(cost_with_tax)

                if not sale_price:
                    sale_price = min_sale_price

                vals.update({
                    "average_cost": average_cost,
                    "tax_amount": tax_amount,
                    "cost_with_tax": cost_with_tax,
                    "min_sale_price": min_sale_price,   # VENTA 30%
                    "sale_price": sale_price,           # VENTA FULL
                    "max_sale_price": sale_price,       # VENTA FULL también
                })

                _logger.info(
                    "[MAC Price Import] Fila %s | Nueva=%s | Anterior=%s | Marca=%s | "
                    "Promedio=%s | IVA=%s | Costo+IVA=%s | Min=%s | Venta=%s",
                    row_number,
                    new_reference,
                    old_reference,
                    brand_line,
                    average_cost,
                    tax_amount,
                    cost_with_tax,
                    min_sale_price,
                    sale_price,
                )

            options.write(vals)

            updated_references += 1
            updated_options += len(options)

        not_found_refs = sorted(set(not_found_refs))

        self.summary = _(
            "Importación de precios finalizada.\n\n"
            "Filas leídas: %(total_rows)s\n"
            "Referencias actualizadas: %(updated_references)s\n"
            "Opciones actualizadas: %(updated_options)s\n"
            "Referencias no encontradas: %(not_found)s\n"
            "Filas omitidas: %(skipped)s\n\n"
            "No encontradas:\n%(not_found_refs)s"
        ) % {
            "total_rows": total_rows,
            "updated_references": updated_references,
            "updated_options": updated_options,
            "not_found": not_found,
            "skipped": skipped,
            "not_found_refs": "\n".join(not_found_refs[:80]),
        }

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Importación finalizada" if updated_options else "Importación de precios"),
                "message": _(
                    "Filas leídas: %(total_rows)s\n"
                    "Referencias actualizadas: %(updated_references)s\n"
                    "Opciones actualizadas: %(updated_options)s\n"
                    "Referencias no encontradas: %(not_found)s\n"
                    "Filas omitidas: %(skipped)s"
                ) % {
                    "total_rows": total_rows,
                    "updated_references": updated_references,
                    "updated_options": updated_options,
                    "not_found": not_found,
                    "skipped": skipped,
                },
                "type": "success" if updated_options else "warning",
                "sticky": not bool(updated_options),
                "next": {"type": "ir.actions.act_window_close"},
            },
        }

    def _read_file(self):
        file_content = base64.b64decode(self.file or b"")
        filename = (self.filename or "").lower()

        if filename.endswith(".xlsx"):
            return self._read_xlsx(file_content)

        if filename.endswith(".csv"):
            return self._read_csv(file_content)

        raise UserError(_("Formato no soportado. Usa archivo .xlsx o .csv."))

    def _read_xlsx(self, file_content):
        try:
            from openpyxl import load_workbook  #type:ignore
        except ImportError as exc:
            raise UserError(_("Debes instalar openpyxl para importar XLSX.")) from exc

        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )

        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)

        try:
            raw_headers = next(iterator)
        except StopIteration:
            return []

        headers = [
            self._normalize_header(header) or f"column_{index}"
            for index, header in enumerate(raw_headers)
        ]

        rows = []
        for row_number, raw_row in enumerate(iterator, start=2):
            row = {}
            for index, value in enumerate(raw_row):
                if index < len(headers):
                    row[headers[index]] = self._cell_to_str(value)

            if any(row.values()):
                rows.append((row_number, row))

        return rows

    def _read_csv(self, file_content):
        text = None

        for encoding in ("utf-8-sig", "latin-1"):
            try:
                text = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            raise UserError(_("No fue posible leer el CSV. Revisa la codificación."))

        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        headers = {
            original: self._normalize_header(original)
            for original in reader.fieldnames or []
        }

        rows = []
        for row_number, raw_row in enumerate(reader, start=2):
            row = {}

            for original_key, value in raw_row.items():
                normalized_key = headers.get(original_key)
                if normalized_key:
                    row[normalized_key] = self._cell_to_str(value)

            if any(row.values()):
                rows.append((row_number, row))

        return rows

    def _get_min_sale_price_from_file(self, row):
        return self._parse_money(
            self._get_value(row, [
                "venta_30",
                "venta_30_",
                "venta_30_porcentaje",
                "venta_30_percent",
                "precio_minimo",
                "min_sale_price",
            ])
        )


    def _get_full_sale_price_from_file(self, row):
        return self._parse_money(
            self._get_value(row, [
                "venta_full",
                "venta_full_",
                "venta_final",
                "precio_full",
                "precio_venta_full",
                "precio_venta",
                "sale_price",
                "precio_publico",
                "precio_al_publico",
            ])
        )

    def _calculate_sale_price(self, cost_with_tax):
        if not cost_with_tax:
            return 0.0

        return round(cost_with_tax / 0.70, 0)

    def _find_options_for_price_import(
        self,
        Option,
        new_reference: str,
        old_reference: str,
        brand_line: str,
    ):
        domains = self._build_price_import_search_domains(
            Option,
            new_reference,
            old_reference,
            brand_line,
        )

        for domain in domains:
            options = Option.search(domain)
            if options:
                _logger.info(
                    "[MAC Price Import] Match | nueva=%s anterior=%s marca=%s domain=%s ids=%s",
                    new_reference,
                    old_reference,
                    brand_line,
                    domain,
                    options.ids,
                )
                return options

        _logger.warning(
            "[MAC Price Import] Sin match | nueva=%s anterior=%s marca=%s",
            new_reference,
            old_reference,
            brand_line,
        )

        return Option.browse()

    def _build_price_import_search_domains(
        self,
        Option,
        new_reference: str,
        old_reference: str,
        brand_line: str,
    ):
        domains = []

        if old_reference and brand_line:
            domains.append([
                ("reference", "=ilike", old_reference),
                "|",
                ("battery_line", "ilike", brand_line),
                ("description", "ilike", brand_line),
            ])

        if old_reference:
            domains.append([("reference", "=ilike", old_reference)])

        if "old_reference" in Option._fields and old_reference:
            domains.append([("old_reference", "=ilike", old_reference)])

        if new_reference and brand_line:
            domains.append([
                ("reference", "=ilike", new_reference),
                "|",
                ("battery_line", "ilike", brand_line),
                ("description", "ilike", brand_line),
            ])

        if new_reference:
            domains.append([("reference", "=ilike", new_reference)])

        return domains

    def _extract_reference(self, description):
        description = self._cell_to_str(description)

        if not description:
            return ""

        match = re.search(r"^\s*([A-Za-z0-9\-]+)", description)
        return match.group(1).strip().upper() if match else ""

    def _get_value(self, row, keys):
        for key in keys:
            value = row.get(self._normalize_header(key))
            if value not in (None, ""):
                return self._cell_to_str(value)

        return ""

    def _parse_number(self, value):
        return self._parse_money(value)

    def _parse_money(self, value):
        value = self._cell_to_str(value)

        if not value:
            return 0.0

        value = (
            value.replace("$", "")
            .replace("COP", "")
            .replace(" ", "")
            .strip()
        )

        if "," in value:
            value = value.replace(".", "").replace(",", ".")
        else:
            parts = value.split(".")
            if len(parts) > 1 and len(parts[-1]) == 3:
                value = value.replace(".", "")

        try:
            return float(value)
        except ValueError:
            return 0.0

    def _normalize_reference(self, value):
        value = self._cell_to_str(value)

        if not value:
            return ""

        return re.sub(r"\s+", "", value.strip().upper())

    def _normalize_brand_line(self, value):
        value = self._cell_to_str(value)

        if not value:
            return ""

        value = self._strip_accents(value).lower()
        value = re.sub(r"[^a-z0-9]+", " ", value)
        value = re.sub(r"\s+", " ", value)

        return value.strip()

    def _cell_to_str(self, value):
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    def _normalize_header(self, value):
        value = self._cell_to_str(value)
        value = self._strip_accents(value)
        value = value.lower()
        value = value.replace("%", "")
        value = value.replace("+", " mas ")
        value = re.sub(r"[^a-z0-9]+", "_", value)
        value = re.sub(r"_+", "_", value)

        return value.strip("_")

    def _strip_accents(self, value):
        value = self._cell_to_str(value)

        return "".join(
            char
            for char in unicodedata.normalize("NFKD", value)
            if not unicodedata.combining(char)
        )
