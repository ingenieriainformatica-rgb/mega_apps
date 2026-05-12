import base64
import csv
import io
import json
import re
import unicodedata
import logging
from typing import Any

from odoo import fields, models, _  # type: ignore
from odoo.exceptions import UserError  # type: ignore

_logger = logging.getLogger(__name__)


class MegaBatteryCatalogImportWizard(models.TransientModel):
    _name = "mega.battery.catalog.import.wizard"
    _description = "Wizard importación catálogo MAC"

    file = fields.Binary(
        string="Archivo",
        required=True,
    )

    filename = fields.Char(
        string="Nombre del archivo",
    )

    publish_on_website = fields.Boolean(
        string="Activar marcas/modelos en website",
        default=True,
        help=(
            "Si está activo, las marcas y modelos creados o usados en la importación "
            "quedarán visibles en el sitio web."
        ),
    )

    application_type = fields.Selection(
        selection=[
            ("light", "Livianos"),
            ("heavy", "Pesados"),
            ("agricultural", "Agrícola"),
            ("industrial", "Industrial / maquinaria"),
            ("other", "Otros"),
        ],
        string="Tipo de aplicación",
        required=True,
        default="light",
    )

    create_missing_brands = fields.Boolean(
        string="Crear marcas faltantes",
        default=True,
    )

    create_missing_models = fields.Boolean(
        string="Crear modelos faltantes",
        default=True,
    )

    link_products = fields.Boolean(
        string="Relacionar productos existentes",
        default=False,
        help=(
            "Busca productos por referencia en nombre o referencia interna. "
            "Para la primera carga del catálogo MAC se recomienda dejarlo apagado."
        ),
    )

    update_existing = fields.Boolean(
        string="Actualizar aplicaciones existentes",
        default=True,
    )

    state = fields.Selection(
        selection=[
            ("draft", "Borrador"),
            ("validated", "Validado"),
            ("done", "Importado"),
        ],
        string="Estado",
        default="draft",
    )

    summary = fields.Text(
        string="Resumen",
        readonly=True,
    )

    line_ids = fields.One2many(
        comodel_name="mega.battery.catalog.import.line",
        inverse_name="wizard_id",
        string="Líneas de validación",
    )

    OPTION_COLUMNS = {
        "mac_nuevas_opc_1": ("mac_new", 1),
        "mac_nuevas_opc_2": ("mac_new", 2),

        "mac_opcion_1": ("mac", 1),
        "mac_opcion_2": ("mac", 2),
        "mac_opcion_3": ("mac", 3),
        "mac_opcion_12_meses": ("mac_12", 1),
        "mac_12_meses": ("mac_12", 1),

        "mac_gold_opc_1": ("mac_gold", 1),
        "mac_gold_opc_2": ("mac_gold", 2),
        "mac_gold_opc_3": ("mac_gold", 3),

        "mac_agm_opc_1": ("mac_agm", 1),
        "mac_agm_opc_2": ("mac_agm", 2),

        "power_taxi_opc_1": ("power_taxi", 1),
        "power_taxi_opc_2": ("power_taxi", 2),
        "power_taxi_opc_3": ("power_taxi", 3),

        "optima_opcion_1": ("optima", 1),
        "optima_opcion_2": ("optima", 2),
        "optima_opc_1": ("optima", 1),
        "optima_opc_2": ("optima", 2),

        "silver_cast": ("silver_cast", 1),

        "coexito_opcion_1": ("coexito", 1),
        "coexito_opcion_2": ("coexito", 2),
        "coexito_opcion_3": ("coexito", 3),
        "coexito_opc_1": ("coexito", 1),
        "coexito_opc_2": ("coexito", 2),
        "coexito_opc_3": ("coexito", 3),
    }

    # ============================================================
    # ACCIONES
    # ============================================================

    def action_validate(self):
        self.ensure_one()

        self.line_ids.unlink()

        rows = self._read_file()
        if not rows:
            raise UserError(_("El archivo no contiene filas para importar."))

        total = 0
        ok = 0
        warnings = 0
        errors = 0

        ImportLine = self.env["mega.battery.catalog.import.line"].sudo()

        for row_number, row in rows:
            total += 1

            values = self._prepare_line_values(row_number, row)

            _logger.info("[MAC Catalog] Línea validada: %s", values)

            if values["status"] == "ok":
                ok += 1
            elif values["status"] == "warning":
                warnings += 1
            else:
                errors += 1

            ImportLine.create(values)

        self.state = "validated"
        self.summary = _(
            "Validación finalizada.\n\n"
            "Filas leídas: %(total)s\n"
            "Correctas: %(ok)s\n"
            "Con advertencias: %(warnings)s\n"
            "Con errores: %(errors)s"
        ) % {
            "total": total,
            "ok": ok,
            "warnings": warnings,
            "errors": errors,
        }

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    def action_import(self):
        self.ensure_one()

        if not self.line_ids:
            self.action_validate()

        importable_lines = self.line_ids.filtered(lambda line: line.status != "error")
        if not importable_lines:
            raise UserError(_("No hay líneas válidas para importar."))

        created = 0
        updated = 0
        skipped = 0

        for line in importable_lines:
            result = self._import_validated_line(line)

            if result == "created":
                created += 1
            elif result == "updated":
                updated += 1
            else:
                skipped += 1

        self.state = "done"

        message = _(
            "Importación finalizada correctamente.\n\n"
            "Aplicaciones creadas: %(created)s\n"
            "Aplicaciones actualizadas: %(updated)s\n"
            "Líneas omitidas: %(skipped)s"
        ) % {
            "created": created,
            "updated": updated,
            "skipped": skipped,
        }

        self.summary = message

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Catálogo MAC importado"),
                "message": message,
                "type": "success",
                "sticky": False,
                "next": {
                    "type": "ir.actions.act_window_close",
                },
            },
        }

    # ============================================================
    # VALIDACIÓN / PREPARACIÓN
    # ============================================================

    def _prepare_line_values(self, row_number, row):
        self.ensure_one()

        brand_name = self._get_value(row, ["marca"])

        original_vehicle_name = self._get_value(
            row,
            [
                "modelo_vehiculo",
                "modelo",
                "modelo_vehiculo_nro_bat",
            ],
        )

        battery_qty = self._parse_int(
            self._get_value(row, ["nro_bat", "nro_bateria", "nro_baterias"])
        ) or 1

        parsed = self._parse_vehicle_name(original_vehicle_name)
        options = self._collect_options(row)

        status = "ok"
        messages = []

        if not brand_name:
            status = "error"
            messages.append("No tiene marca.")

        if not original_vehicle_name:
            status = "error"
            messages.append("No tiene modelo / vehículo.")

        if not options:
            status = "error"
            messages.append("No tiene referencias de batería.")

        if parsed["year_from"] and parsed["year_to"]:
            if parsed["year_from"] > parsed["year_to"]:
                status = "error"
                messages.append("El año desde no puede ser mayor que el año hasta.")

        brand = False

        if brand_name and status != "error":
            brand = self._find_brand(brand_name)

            if not brand:
                if self.create_missing_brands:
                    if self.publish_on_website:
                        messages.append("La marca será creada y quedará activa en website.")
                    else:
                        messages.append("La marca será creada.")
                else:
                    status = "error"
                    messages.append("La marca no existe en Flota.")

        if parsed["model_base_name"] and status != "error":
            if brand:
                model = self._find_model(brand, parsed["model_base_name"])

                if not model:
                    if self.create_missing_models:
                        if self.publish_on_website:
                            messages.append("El modelo será creado y quedará activo en website.")
                        else:
                            messages.append("El modelo será creado.")
                    else:
                        status = "error"
                        messages.append("El modelo no existe en Flota.")
            else:
                if self.create_missing_brands and self.create_missing_models:
                    if self.publish_on_website:
                        messages.append(
                            "El modelo será creado después de crear la marca y quedará activo en website."
                        )
                    else:
                        messages.append("El modelo será creado después de crear la marca.")
                elif self.create_missing_brands and not self.create_missing_models:
                    status = "error"
                    messages.append(
                        "La marca será creada, pero el modelo no se puede crear porque la opción está desactivada."
                    )

        missing_products = []
        if self.link_products and options:
            for option in options:
                product = self._find_product_by_reference(option["reference"])
                if not product:
                    missing_products.append(option["reference"])

        if missing_products:
            # En esta fase NO es warning. Se guardan referencias como texto.
            messages.append(
                "Referencias sin producto relacionado; se guardarán como texto: %s"
                % ", ".join(sorted(set(missing_products)))
            )

        return {
            "wizard_id": self.id,
            "row_number": row_number,
            "brand_name": brand_name,
            "model_base_name": parsed["model_base_name"],
            "original_vehicle_name": original_vehicle_name,
            "fuel_type": parsed["fuel_type"],
            "year_from": parsed["year_from"],
            "year_to": parsed["year_to"],
            "engine_capacity": parsed["engine_capacity"],
            "start_stop": parsed["start_stop"],
            "battery_qty": battery_qty,
            "options_json": json.dumps(options, ensure_ascii=False),
            "status": status,
            "message": "\n".join(messages) if messages else "OK",
        }

    # ============================================================
    # IMPORTACIÓN
    # ============================================================

    def _import_validated_line(self, line):
        self.ensure_one()

        brand = self._get_or_create_brand(line.brand_name)
        if not brand:
            return "skipped"

        model = self._get_or_create_vehicle_model(
            brand=brand,
            model_name=line.model_base_name,
        )
        if not model:
            return "skipped"

        Application = self.env["mega.battery.application"].sudo()

        domain = [
            ("application_type", "=", self.application_type),
            ("brand_id", "=", brand.id),
            ("model_id", "=", model.id),
            ("original_vehicle_name", "=", line.original_vehicle_name),
            ("year_from", "=", line.year_from or 0),
            ("year_to", "=", line.year_to or 0),
            ("engine_capacity", "=", line.engine_capacity or False),
        ]

        application = Application.search(domain, limit=1)

        vals = {
            "application_type": self.application_type,
            "brand_id": brand.id,
            "model_id": model.id,
            "original_vehicle_name": line.original_vehicle_name,
            "fuel_type": line.fuel_type,
            "year_from": line.year_from,
            "year_to": line.year_to,
            "engine_capacity": line.engine_capacity,
            "start_stop": line.start_stop,
            "battery_qty": line.battery_qty or 1,
        }

        if application and not self.update_existing:
            return "skipped"

        if application:
            application.write(vals)
            application.option_ids.unlink()
            result = "updated"
        else:
            application = Application.create(vals)
            result = "created"

        options = json.loads(line.options_json or "[]")
        sequence = 10

        Option = self.env["mega.battery.application.option"].sudo()

        for option in options:
            product = False
            if self.link_products:
                product = self._find_product_by_reference(option["reference"])

            Option.create({
                "application_id": application.id,
                "sequence": sequence,
                "battery_line": option["battery_line"],
                "option_number": option["option_number"],
                "reference": option["reference"],
                "product_id": product.id if product else False,
            })

            sequence += 10

        return result

    # ============================================================
    # WEBSITE / FLEET HELPERS
    # ============================================================

    def _website_visibility_vals(self, model_name: str) -> dict[str, Any]:
        self.ensure_one()

        if not self.publish_on_website:
            return {}

        Model = self.env[model_name].sudo()

        if "show_on_website" in Model._fields:
            return {"show_on_website": True}

        return {}

    def _publish_record_on_website(self, record):
        self.ensure_one()

        if not self.publish_on_website or not record:
            return

        if "show_on_website" in record._fields and not record.show_on_website:
            record.sudo().write({
                "show_on_website": True,
            })

    def _get_or_create_brand(self, brand_name):
        self.ensure_one()

        Brand = self.env["fleet.vehicle.model.brand"].sudo()

        brand_name = self._cell_to_str(brand_name)
        if not brand_name:
            return False

        brand = self._find_brand(brand_name)

        if brand:
            self._publish_record_on_website(brand)
            return brand

        if not self.create_missing_brands:
            return False

        vals = {
            "name": self._format_catalog_name(brand_name),
        }
        vals.update(
            self._website_visibility_vals("fleet.vehicle.model.brand")
        )

        brand = Brand.create(vals)
        self._publish_record_on_website(brand)

        return brand

    def _get_or_create_vehicle_model(self, brand, model_name):
        self.ensure_one()

        VehicleModel = self.env["fleet.vehicle.model"].sudo()

        model_name = self._cell_to_str(model_name)
        if not brand or not model_name:
            return False

        vehicle_model = self._find_model(brand, model_name)

        if vehicle_model:
            self._publish_record_on_website(vehicle_model)
            return vehicle_model

        if not self.create_missing_models:
            return False

        vals = {
            "name": self._format_catalog_name(model_name),
            "brand_id": brand.id,
        }
        vals.update(
            self._website_visibility_vals("fleet.vehicle.model")
        )

        vehicle_model = VehicleModel.create(vals)
        self._publish_record_on_website(vehicle_model)

        return vehicle_model

    # ============================================================
    # LECTURA DE ARCHIVOS
    # ============================================================

    def _read_file(self):
        self.ensure_one()

        file_content = base64.b64decode(self.file or b"")
        filename = (self.filename or "").lower()

        if filename.endswith(".csv"):
            return self._read_csv(file_content)

        if filename.endswith(".xlsx"):
            return self._read_xlsx(file_content)

        raise UserError(_("Formato no soportado. Usa archivo .csv o .xlsx."))

    def _read_csv(self, file_content):
        text = None

        for encoding in ("utf-8-sig", "latin-1"):
            try:
                text = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            raise UserError(_("No fue posible leer el CSV. Revisa la codificación."))

        sample = text[:4096]

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)

        headers = {
            original: self._normalize_header(original)
            for original in reader.fieldnames or []
        }

        rows = []
        for row_number, raw_row in enumerate(reader, start=2):
            row = {}

            for original_key, value in raw_row.items():
                normalized_key = headers.get(original_key)
                if normalized_key:
                    row[normalized_key] = self._cell_to_str(value)

            if any(row.values()):
                rows.append((row_number, row))

        return rows

    def _read_xlsx(self, file_content):
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise UserError(
                _(
                    "Para importar XLSX debes instalar la librería openpyxl "
                    "en el entorno de Odoo."
                )
            ) from exc

        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )

        worksheets = workbook.worksheets

        if not worksheets:
            raise UserError(_("El archivo XLSX no contiene hojas para importar."))

        sheet = worksheets[0]
        iterator = sheet.iter_rows(values_only=True)

        try:
            raw_headers = next(iterator)
        except StopIteration:
            return []

        headers = []
        for index, header in enumerate(raw_headers):
            normalized = self._normalize_header(header)
            headers.append(normalized or f"column_{index}")

        rows = []
        for row_number, raw_row in enumerate(iterator, start=2):
            row = {}

            for index, value in enumerate(raw_row):
                if index >= len(headers):
                    continue

                row[headers[index]] = self._cell_to_str(value)

            if any(row.values()):
                rows.append((row_number, row))

        return rows

    # ============================================================
    # PARSEO CATÁLOGO MAC
    # ============================================================

    def _collect_options(self, row):
        options = []

        for column_name, option_config in self.OPTION_COLUMNS.items():
            value = self._get_value(row, [column_name])
            if not value:
                continue

            battery_line, option_number = option_config

            for reference in self._split_references(value):
                options.append({
                    "battery_line": battery_line,
                    "option_number": option_number,
                    "reference": reference,
                })

        return options

    def _split_references(self, value):
        value = self._cell_to_str(value)
        if not value:
            return []

        raw_parts = re.split(r"[\n;,]+", value)
        references = []

        for part in raw_parts:
            part = part.strip()
            if part:
                references.append(part)

        return references

    def _parse_vehicle_name(self, original_name):
        text = self._cell_to_str(original_name)
        clean_text = re.sub(r"\s+", " ", text).strip()

        year_from = 0
        year_to = 0

        range_match = re.search(
            r"\b(19\d{2}|20\d{2})\s*(?:al|a)\s*(19\d{2}|20\d{2})\b",
            clean_text,
            flags=re.IGNORECASE,
        )

        if range_match:
            year_from = int(range_match.group(1))
            year_to = int(range_match.group(2))
        else:
            single_year_match = re.search(
                r"\b(19\d{2}|20\d{2})\b",
                clean_text,
                flags=re.IGNORECASE,
            )
            if single_year_match:
                year_from = int(single_year_match.group(1))
                year_to = int(single_year_match.group(1))

        engine_capacity = ""
        engine_match = re.search(
            r"\b\d+(?:[.,]\d+)?\s*L\b",
            clean_text,
            flags=re.IGNORECASE,
        )

        if engine_match:
            engine_capacity = engine_match.group(0).replace(" ", "").upper()

        start_stop = bool(
            re.search(
                r"\b(?:start|star)[\s-]*stop\b",
                clean_text,
                flags=re.IGNORECASE,
            )
        )

        fuel_type = ""
        fuel_patterns = [
            "Hibrido Gasolina",
            "Híbrido Gasolina",
            "Hibrido",
            "Híbrido",
            "Gasolina",
            "Diesel",
            "Electrico",
            "Eléctrico",
            "GNV",
            "GLP",
        ]

        fuel_pos = None
        for fuel in fuel_patterns:
            match = re.search(re.escape(fuel), clean_text, flags=re.IGNORECASE)
            if match:
                fuel_type = fuel
                fuel_pos = match.start()
                break

        cut_positions = []

        if fuel_pos is not None:
            cut_positions.append(fuel_pos)

        if range_match:
            cut_positions.append(range_match.start())
        elif year_from:
            single_year_match = re.search(r"\b%s\b" % year_from, clean_text)
            if single_year_match:
                cut_positions.append(single_year_match.start())

        if engine_match:
            cut_positions.append(engine_match.start())

        if cut_positions:
            model_base_name = clean_text[:min(cut_positions)].strip()
        else:
            model_base_name = clean_text

        model_base_name = re.sub(r"\s+", " ", model_base_name).strip(" -/")

        return {
            "model_base_name": model_base_name or clean_text,
            "fuel_type": fuel_type,
            "year_from": year_from,
            "year_to": year_to,
            "engine_capacity": engine_capacity,
            "start_stop": start_stop,
        }

    # ============================================================
    # BÚSQUEDAS
    # ============================================================

    def _find_brand(self, brand_name):
        brand_name = self._cell_to_str(brand_name)
        if not brand_name:
            return False

        Brand = self.env["fleet.vehicle.model.brand"].sudo()

        brands = Brand.search([
            ("name", "ilike", brand_name),
        ])

        normalized_target = self._normalize_text(brand_name)

        for brand in brands:
            if self._normalize_text(brand.name) == normalized_target:
                return brand

        return False

    def _find_model(self, brand, model_name):
        model_name = self._cell_to_str(model_name)
        if not brand or not model_name:
            return False

        VehicleModel = self.env["fleet.vehicle.model"].sudo()

        models = VehicleModel.search([
            ("brand_id", "=", brand.id),
            ("name", "ilike", model_name),
        ])

        normalized_target = self._normalize_text(model_name)

        for model in models:
            if self._normalize_text(model.name) == normalized_target:
                return model

        return False

    def _find_product_by_reference(self, reference):
        reference = self._cell_to_str(reference)
        if not reference:
            return False

        Product = self.env["product.template"].sudo()

        product = Product.search([
            ("default_code", "=ilike", reference),
        ], limit=1)

        if product:
            return product

        return Product.search([
            ("name", "=ilike", reference),
        ], limit=1)

    # ============================================================
    # UTILIDADES
    # ============================================================

    def _get_value(self, row, keys):
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                return self._cell_to_str(value)

        return ""

    def _parse_int(self, value):
        value = self._cell_to_str(value)
        if not value:
            return 0

        try:
            return int(float(value.replace(",", ".")))
        except ValueError:
            return 0

    def _cell_to_str(self, value):
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    def _normalize_header(self, value):
        value = self._cell_to_str(value)
        value = self._strip_accents(value)
        value = value.lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        value = re.sub(r"_+", "_", value)

        return value.strip("_")

    def _normalize_text(self, value):
        value = self._cell_to_str(value)
        value = self._strip_accents(value)
        value = value.upper()
        value = re.sub(r"[^A-Z0-9]+", " ", value)
        value = re.sub(r"\s+", " ", value)

        return value.strip()

    def _strip_accents(self, value):
        value = self._cell_to_str(value)

        return "".join(
            char
            for char in unicodedata.normalize("NFKD", value)
            if not unicodedata.combining(char)
        )

    def _format_catalog_name(self, value):
        value = self._cell_to_str(value)
        if not value:
            return value

        words = value.lower().split()
        formatted_words = []

        for word in words:
            clean_word = re.sub(r"[^a-zA-Z0-9]", "", word)

            if len(clean_word) <= 3 and clean_word:
                formatted_words.append(word.upper())
            else:
                formatted_words.append(word.capitalize())

        return " ".join(formatted_words)


class MegaBatteryCatalogImportLine(models.TransientModel):
    _name = "mega.battery.catalog.import.line"
    _description = "Línea validación importación catálogo MAC"
    _order = "row_number"

    wizard_id = fields.Many2one(
        comodel_name="mega.battery.catalog.import.wizard",
        required=True,
        ondelete="cascade",
    )

    row_number = fields.Integer(
        string="Fila",
    )

    brand_name = fields.Char(
        string="Marca",
    )

    model_base_name = fields.Char(
        string="Modelo base",
    )

    original_vehicle_name = fields.Char(
        string="Modelo / vehículo original",
    )

    fuel_type = fields.Char(
        string="Combustible",
    )

    year_from = fields.Integer(
        string="Año desde",
    )

    year_to = fields.Integer(
        string="Año hasta",
    )

    engine_capacity = fields.Char(
        string="Cilindraje",
    )

    start_stop = fields.Boolean(
        string="Start-Stop",
    )

    battery_qty = fields.Integer(
        string="Nro. baterías",
    )

    options_json = fields.Text(
        string="Opciones JSON",
    )

    status = fields.Selection(
        selection=[
            ("ok", "OK"),
            ("warning", "Advertencia"),
            ("error", "Error"),
        ],
        string="Estado",
        default="ok",
    )

    message = fields.Text(
        string="Mensaje",
    )
