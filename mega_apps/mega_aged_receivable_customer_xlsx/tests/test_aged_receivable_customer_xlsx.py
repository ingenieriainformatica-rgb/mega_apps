# -*- coding: utf-8 -*-
import io

from odoo import Command, fields
from odoo.tests import tagged

from odoo.addons.account_reports.tests.common import TestAccountReportsCommon

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@tagged('post_install', '-at_install')
class TestAgedReceivableCustomerXlsx(TestAccountReportsCommon):
    """ Covers the 'Cliente' column added to the Aged Receivable XLSX export by this module:
    it must be present and correctly repeated on every detail row of the exported file, while the
    interactive report table and the PDF export must remain exactly as the standard report defines them. """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.report = cls.env.ref('account_reports.aged_receivable_report')
        cls.date_to = fields.Date.from_string('2024-06-30')
        cls.old_date = fields.Date.from_string('2023-01-15')

        # A child contact of partner_a: invoices billed to it must show partner_a's name
        # (its commercial_partner_id), not the child contact's own name.
        cls.partner_a_child = cls.env['res.partner'].create({
            'name': 'Partner A - Branch Office',
            'parent_id': cls.partner_a.id,
        })
        cls.partner_a.vat = '900123456-1'

    # -------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------

    def _invoice(self, partner, price_unit, invoice_date, company=None, post=True):
        invoice = self._create_invoice_one_line(
            move_type='out_invoice',
            partner_id=partner.id,
            price_unit=price_unit,
            invoice_date=invoice_date,
            tax_ids=[],
            company_id=(company or self.company_data['company']).id,
        )
        if post:
            invoice.action_post()
        return invoice

    def _get_screen_options(self, partner_ids=None, companies=None):
        # unfold_all=True mirrors a user expanding every customer before downloading (as instructed
        # in the manual test steps): without it, the export only contains the folded partner/total
        # rows and no per-document detail rows at all.
        # 'forced_companies' (rather than a context trick) is what survives the internal
        # get_options() re-derivation that export_to_xlsx does on its way to the XLSX file.
        default_options = {'unfold_all': True}
        if companies is not None:
            default_options['forced_companies'] = companies.ids
        options = self._generate_options(self.report, self.date_to, self.date_to, default_options=default_options)
        if partner_ids is not None:
            options['partner_ids'] = partner_ids
        return options

    def _export_xlsx_rows(self, options):
        # The 'Cliente' column only ever comes from the dedicated cog-menu action, never from the
        # standard XLSX button (see models/account_report.py) - this is what the user actually clicks
        # after opening the report's gear/cog menu.
        file_data = self.report.mega_export_aged_receivable_customer_xlsx(options)
        self.assertEqual(file_data['file_type'], 'xlsx')
        if load_workbook is None:
            self.skipTest("openpyxl is not available")
        workbook = load_workbook(filename=io.BytesIO(file_data['file_content']), data_only=True)
        sheet = workbook.worksheets[0]
        return list(sheet.rows), sheet

    def _find_header_row(self, rows):
        for row in rows:
            values = [cell.value for cell in row]
            if 'Cliente' in values:
                return values
        self.fail("The 'Cliente' header was not found in the exported XLSX file.")

    # -------------------------------------------------------------------
    # Data-content tests
    # -------------------------------------------------------------------

    def test_partner_name_header_is_not_duplicated(self):
        """ Only the leading custom 'Cliente' column should be exported, even if an old duplicate
        report column with the same expression label still exists in the database. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)

        self.assertEqual(header.count('Cliente'), 1)

    def test_partner_name_repeated_for_several_invoices_same_customer(self):
        """ Scenarios 1 & 9: one customer with several pending invoices; the client column
        must be repeated on every one of its detail rows, and per-customer totals stay correct. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        self._invoice(self.partner_a, 250.0, self.old_date)
        self._invoice(self.partner_a, 50.0, self.date_to)

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        detail_rows = [
            [cell.value for cell in row]
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        ]
        # The three invoices must each produce a detail row carrying the customer's name.
        self.assertEqual(len(detail_rows), 3)

    def test_partner_name_two_different_customers(self):
        """ Scenario 2: two different customers, each detail row must show its own customer, never the other's. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        self._invoice(self.partner_b, 200.0, self.old_date)

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        names_in_file = {
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value not in (None, 'Cliente')
        }
        self.assertIn(self.partner_a.name, names_in_file)
        self.assertIn(self.partner_b.name, names_in_file)

    def test_partner_name_partially_paid_invoice(self):
        """ Scenario 3: a partially paid invoice must still carry the customer name on its remaining line. """
        invoice = self._invoice(self.partner_a, 300.0, self.old_date)
        self._register_payment(invoice, amount=100.0)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        detail_names = [
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        ]
        self.assertTrue(detail_names, "The partially paid invoice's remaining balance should still show the customer.")

    def test_partner_name_credit_note(self):
        """ Scenario 4: a credit note against an invoice must also carry the customer name.
        Issued standalone (not reversing/auto-reconciling the invoice) so both remain open,
        distinct receivable-account lines, which is what this scenario is actually about. """
        self._invoice(self.partner_a, 400.0, self.old_date)
        credit_note = self._create_invoice_one_line(
            move_type='out_refund',
            partner_id=self.partner_a.id,
            price_unit=100.0,
            invoice_date=self.old_date,
            tax_ids=[],
            company_id=self.company_data['company'].id,
        )
        credit_note.action_post()

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        detail_rows_count = sum(
            1 for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        )
        # Both the invoice and its credit note must appear as separate detail rows.
        self.assertGreaterEqual(detail_rows_count, 2)

    def test_partner_name_unreconciled_payment(self):
        """ Scenario 5: an unreconciled customer payment (advance) posted on the receivable
        account must also show the payer's name on its own detail row. """
        payment = self.env['account.payment'].create({
            'payment_type': 'inbound',
            'partner_type': 'customer',
            'partner_id': self.partner_a.id,
            'amount': 75.0,
            'date': self.old_date,
            'journal_id': self.company_data['default_journal_bank'].id,
        })
        payment.action_post()
        self.assertFalse(payment.move_id.line_ids.filtered(
            lambda l: l.account_id.account_type == 'asset_receivable'
        ).reconciled)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        detail_names = [
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        ]
        self.assertTrue(detail_names, "The unreconciled payment should show the customer on its own row.")

    def test_document_date_filled_for_payment_entries(self):
        """ 'Fecha de factura' (invoice_date) is a standard column that stays blank for anything
        that isn't an invoice, e.g. a manual customer payment - which is exactly the gap that was
        reported. 'Fecha del documento' must instead show that payment's own accounting date. """
        payment = self.env['account.payment'].create({
            'payment_type': 'inbound',
            'partner_type': 'customer',
            'partner_id': self.partner_a.id,
            'amount': 75.0,
            'date': self.old_date,
            'journal_id': self.company_data['default_journal_bank'].id,
        })
        payment.action_post()

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')
        invoice_date_idx = header.index('Invoice Date')
        document_date_idx = header.index('Fecha del documento')

        detail_row = next(
            row for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        )
        self.assertIn(detail_row[invoice_date_idx].value, (None, ''), "Payments have no invoice_date, by design.")
        # Exact date-vs-string round-tripping through the xlsx cell is a pre-existing, locale-dependent
        # quirk of the core exporter (_get_cell_type_value) shared by every 'date' figure_type column,
        # including the standard 'Invoice Date' one - not something specific to this new column. What
        # matters here is that it's no longer blank, unlike 'Invoice Date' on this same row.
        self.assertNotIn(detail_row[document_date_idx].value, (None, ''), "This is exactly the gap that was reported: payments had no date at all.")

    def test_document_date_filled_for_regular_invoices_too(self):
        """ For an actual invoice, 'Fecha del documento' must be filled too - it isn't only meant
        for payments, every detail row gets it. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')
        document_date_idx = header.index('Fecha del documento')

        detail_row = next(
            row for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        )
        self.assertNotIn(detail_row[document_date_idx].value, (None, ''))

    def test_document_date_absent_from_screen_pdf_and_standard_button(self):
        """ 'Fecha del documento', like 'Cliente'/'NIT', must only ever appear in the dedicated
        cog-menu export. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        expression_labels = {column['expression_label'] for column in options['columns']}
        self.assertNotIn('document_date', expression_labels)

        print_options = self.report.get_options(previous_options={**options, 'export_mode': 'print'})
        self.assertNotIn('document_date', {c['expression_label'] for c in print_options['columns']})

        file_data = self.report.export_to_xlsx(options)
        if load_workbook is None:
            self.skipTest("openpyxl is not available")
        workbook = load_workbook(filename=io.BytesIO(file_data['file_content']), data_only=True)
        header_values = [cell.value for row in workbook.worksheets[0].rows for cell in row]
        self.assertNotIn('Fecha del documento', header_values)

    def test_partner_name_blank_without_partner(self):
        """ Scenario 6: a receivable-account journal item without a partner must not invent a value. """
        misc_move = self.env['account.move'].create({
            'date': self.old_date,
            'line_ids': [
                Command.create({'debit': 500.0, 'credit': 0.0, 'account_id': self.company_data['default_account_receivable'].id}),
                Command.create({'debit': 0.0, 'credit': 500.0, 'account_id': self.company_data['default_account_revenue'].id}),
            ],
        })
        misc_move.action_post()

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        # The 'Unknown' group row itself is expected to be blank regardless (group rows never carry
        # a customer name); what matters is the *detail* row right below it, for the actual journal
        # item, which must also stay blank instead of inventing a value.
        unknown_row_index = next(
            index for index, row in enumerate(rows)
            if row[0].value == 'Unknown'
        )
        detail_row = rows[unknown_row_index + 1]
        self.assertNotEqual(detail_row[0].value, 'Unknown', "Expected the detail row following the 'Unknown' group.")
        self.assertIn(detail_row[cliente_idx].value, (None, ''))

    def test_partner_name_follows_invoice_partner_not_stale_move_line_partner(self):
        """ Regression: a real database was found where a journal item's own partner_id had drifted
        away from its invoice's partner_id (the invoice form showed a different customer than the
        one the report's own grouping/labelling used, which reads account.move.line.partner_id).
        'Cliente' must reflect the document's actual customer (account.move.partner_id, what opening
        the invoice shows), not whatever partner_id happens to sit on the journal line itself. """
        invoice = self._invoice(self.partner_a, 100.0, self.old_date)
        receivable_line = invoice.line_ids.filtered(lambda l: l.account_id.account_type == 'asset_receivable')
        # Force the exact real-world inconsistency directly at the SQL level (this isn't reachable
        # through the normal ORM/UI flow, which is exactly why it went unnoticed on that database).
        self.env.cr.execute(
            "UPDATE account_move_line SET partner_id = %s WHERE id = %s",
            (self.partner_b.id, receivable_line.id),
        )
        receivable_line.invalidate_recordset(['partner_id'])

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        detail_row = next(row for row in rows if row[0].value == invoice.name)
        self.assertEqual(
            detail_row[cliente_idx].value, self.partner_a.name,
            "Cliente must follow the invoice's own partner (account.move.partner_id), "
            "not a stale/divergent account.move.line.partner_id.",
        )

    def test_partner_name_uses_commercial_partner_for_child_contact(self):
        """ Scenario 7: invoicing a child contact must show the parent commercial partner's name. """
        self._invoice(self.partner_a_child, 150.0, self.old_date)

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        names_in_file = {
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value not in (None, '', 'Cliente')
        }
        self.assertIn(self.partner_a.name, names_in_file)
        self.assertNotIn(self.partner_a_child.name, names_in_file)

    def test_partner_vat_repeated_per_row(self):
        """ The 'NIT' column must repeat the customer's tax id (commercial_partner_id.vat) on every
        detail row, right next to 'Cliente', the same way and for the same reasons. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        self._invoice(self.partner_a, 50.0, self.old_date)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')
        vat_idx = header.index('NIT')

        detail_rows = [
            row for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        ]
        self.assertEqual(len(detail_rows), 2)
        for row in detail_rows:
            self.assertEqual(row[vat_idx].value, self.partner_a.vat)

    def test_partner_vat_uses_commercial_partner_for_child_contact(self):
        """ Same as Cliente: a child contact's invoice must show the parent commercial partner's NIT. """
        self._invoice(self.partner_a_child, 150.0, self.old_date)

        options = self._get_screen_options()
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')
        vat_idx = header.index('NIT')

        detail_row = next(row for row in rows if row[cliente_idx].value == self.partner_a.name)
        self.assertEqual(detail_row[vat_idx].value, self.partner_a.vat)

    def test_partner_vat_blank_when_partner_has_no_vat(self):
        """ A customer without a tax id must leave the NIT cell blank, never inventing a value. """
        self._invoice(self.partner_b, 200.0, self.old_date)
        self.assertFalse(self.partner_b.vat)

        options = self._get_screen_options(partner_ids=self.partner_b.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')
        vat_idx = header.index('NIT')

        detail_row = next(row for row in rows if row[cliente_idx].value == self.partner_b.name)
        self.assertIn(detail_row[vat_idx].value, (None, ''))

    def test_partner_vat_absent_from_screen_pdf_and_standard_button(self):
        """ 'NIT', like 'Cliente', must only ever appear in the dedicated cog-menu export. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        expression_labels = {column['expression_label'] for column in options['columns']}
        self.assertNotIn('partner_vat', expression_labels)

        print_options = self.report.get_options(previous_options={**options, 'export_mode': 'print'})
        self.assertNotIn('partner_vat', {c['expression_label'] for c in print_options['columns']})

        file_data = self.report.export_to_xlsx(options)
        if load_workbook is None:
            self.skipTest("openpyxl is not available")
        workbook = load_workbook(filename=io.BytesIO(file_data['file_content']), data_only=True)
        header_values = [cell.value for row in workbook.worksheets[0].rows for cell in row]
        self.assertNotIn('NIT', header_values)

    def test_partner_name_numeric_cells_stay_numeric(self):
        """ Scenario 16: amounts in the exported file must remain numbers, not text, regardless of the new column. """
        self._invoice(self.partner_a, 123.45, self.old_date)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        # Not every column past 'Cliente' is monetary (e.g. 'Fecha de factura' is a date, and
        # 'Cuenta'/'Moneda' show up as text because _generate_options sets show_account/show_currency),
        # so this only asserts that the period-amount cells - which are numeric on every detail row,
        # even when 0.0 (see core test_aged_receivable_report.py) - were not turned into text.
        checked_a_row = False
        for row in rows:
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name:
                numeric_cells = [cell for cell in row[cliente_idx + 1:] if isinstance(cell.value, (int, float))]
                self.assertTrue(numeric_cells, "Expected at least one numeric amount cell on this detail row.")
                checked_a_row = True
        self.assertTrue(checked_a_row, "Expected at least one detail row for partner_a.")

    def test_partner_name_totals_match_between_odoo_and_xlsx(self):
        """ Scenario 9/19: XLSX totals must equal the report's own totals; the extra column
        must not duplicate or alter any amount. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        self._invoice(self.partner_a, 250.0, self.old_date)
        self._invoice(self.partner_b, 400.0, self.old_date)

        options = self._get_screen_options()
        lines = self.report._get_lines(options)
        # lines[0] is always the report's own root line ('Aged Receivable'), i.e. the grand total
        # across every customer; its last column is 'total' (last one declared on the report).
        root_line = lines[0]
        self.assertEqual(root_line['name'], 'Aged Receivable')
        odoo_total = root_line['columns'][-1]['no_format']

        _rows, sheet = self._export_xlsx_rows(options)
        # Header rows can wrap; search all rows for the one containing 'Total' as the last populated header.
        total_col_idx = None
        for row in sheet.rows:
            values = [c.value for c in row]
            if 'Total' in values:
                total_col_idx = values.index('Total')
                break
        self.assertIsNotNone(total_col_idx, "Could not locate the 'Total' column header in the XLSX file.")

        xlsx_grand_total = None
        for row in sheet.rows:
            if row[0].value == 'Aged Receivable' and len(row) > total_col_idx:
                xlsx_grand_total = row[total_col_idx].value
                break
        self.assertIsNotNone(xlsx_grand_total)
        self.assertAlmostEqual(xlsx_grand_total, odoo_total, places=2)

    # -------------------------------------------------------------------
    # Filters
    # -------------------------------------------------------------------

    def test_partner_name_respects_partner_filter(self):
        """ Scenario 12: filtering the report by customer must also restrict the XLSX 'Cliente' values. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        self._invoice(self.partner_b, 200.0, self.old_date)

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        names_in_file = {
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value not in (None, '', 'Cliente')
        }
        self.assertIn(self.partner_a.name, names_in_file)
        self.assertNotIn(self.partner_b.name, names_in_file)

    def test_partner_name_respects_date_filter(self):
        """ Scenario 11: an invoice dated after the report's cut-off date must not appear at all. """
        self._invoice(self.partner_a, 100.0, self.old_date)
        future_invoice = self._invoice(self.partner_a, 999.0, fields.Date.from_string('2025-01-01'))

        options = self._get_screen_options(partner_ids=self.partner_a.ids)
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        doc_names = {
            row[0].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value == self.partner_a.name
        }
        self.assertNotIn(future_invoice.name, doc_names)

    def test_partner_name_respects_multi_company_visibility(self):
        """ Scenario 20: a customer belonging to another company must never leak into the export
        when that company is not part of the currently allowed/selected companies. """
        other_partner = self.env['res.partner'].create({'name': 'Other Company Customer'})
        self._invoice(self.partner_a, 100.0, self.old_date, company=self.company_data['company'])
        self._invoice(other_partner, 300.0, self.old_date, company=self.company_data_2['company'])

        options = self._get_screen_options(companies=self.company_data['company'])
        rows, _sheet = self._export_xlsx_rows(options)
        header = self._find_header_row(rows)
        cliente_idx = header.index('Cliente')

        names_in_file = {
            row[cliente_idx].value
            for row in rows
            if len(row) > cliente_idx and row[cliente_idx].value not in (None, '', 'Cliente')
        }
        self.assertIn(self.partner_a.name, names_in_file)
        self.assertNotIn(other_partner.name, names_in_file)

    # -------------------------------------------------------------------
    # Scope: only the dedicated cog-menu action is affected
    # -------------------------------------------------------------------

    def test_cog_menu_has_dedicated_download_entry(self):
        """ The report's gear/cog dropdown (where 'Copy to Documents' lives) must offer a new,
        separate entry for this export - the standard XLSX toolbar button must not be touched at all. """
        options = self._get_screen_options()

        custom_buttons = [b for b in options['buttons'] if b.get('action_param') == 'mega_export_aged_receivable_customer_xlsx']
        self.assertEqual(len(custom_buttons), 1)
        self.assertFalse(custom_buttons[0].get('always_show'), "Must stay in the cog dropdown, not the toolbar.")

        standard_xlsx_button = next(b for b in options['buttons'] if b.get('action_param') == 'export_to_xlsx')
        self.assertTrue(standard_xlsx_button.get('always_show'), "The standard XLSX toolbar button must be untouched.")

    def test_partner_name_column_absent_from_screen_report(self):
        """ Scenario 18: the interactive report must be unaffected: no 'partner_name' column, no crash. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        expression_labels = {column['expression_label'] for column in options['columns']}
        self.assertNotIn('partner_name', expression_labels)

        # The report must still render normally (no exception, totals computed as usual).
        lines = self.report._get_lines(options)
        self.assertTrue(lines)

    def test_partner_name_column_absent_from_standard_xlsx_button(self):
        """ The plain 'XLSX' toolbar button (export_to_xlsx) must never include 'Cliente': only the
        new cog-menu entry (mega_export_aged_receivable_customer_xlsx) does. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        file_data = self.report.export_to_xlsx(options)
        workbook = load_workbook(filename=io.BytesIO(file_data['file_content']), data_only=True) if load_workbook else None
        if workbook is None:
            self.skipTest("openpyxl is not available")
        header_values = [cell.value for row in workbook.worksheets[0].rows for cell in row]
        self.assertNotIn('Cliente', header_values)

    def test_partner_name_column_absent_from_pdf_export(self):
        """ Scenario 17: the PDF export path must render exactly as before, without the extra column. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        print_options = self.report.get_options(previous_options={**options, 'export_mode': 'print'})
        expression_labels = {column['expression_label'] for column in print_options['columns']}
        self.assertNotIn('partner_name', expression_labels)

        # Renders the same HTML body used to feed wkhtmltopdf, without requiring the binary itself.
        lines = self.report._filter_out_folded_children(self.report._get_lines(print_options))
        html = self.report._get_pdf_export_html(print_options, lines)
        self.assertNotIn('Cliente', html)

    def test_partner_name_does_not_leak_to_screen_after_download(self):
        """ Regression test: an earlier version stored the 'show partner name' flag on `options`
        itself, which the report's on-screen state echoes back and reuses on every later request -
        so the column stayed visible on screen forever after a single download. Calling the dedicated
        export must not affect any *later, independent* call, screen or otherwise. """
        self._invoice(self.partner_a, 100.0, self.old_date)

        options = self._get_screen_options()
        self.report.mega_export_aged_receivable_customer_xlsx(options)

        # A brand new options round-trip (as a screen refresh, unfold, or filter change would do)
        # must come back clean, with no trace of the previous download.
        fresh_options = self.report.get_options(previous_options=options)
        expression_labels = {column['expression_label'] for column in fresh_options['columns']}
        self.assertNotIn('partner_name', expression_labels)
